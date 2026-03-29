# EconBot_v109 — Clean rebuild (guild-only commands, legacy DB partition, bank message IDs persisted in Postgres)
# NOTE: This is a full replacement for main.py (Railway runs /app/main.py).
# Constraints honored:
# - Character-based economy (not user-based)
# - No invented env vars (uses only the approved set)
# - No invented DB tables EXCEPT the explicitly approved bank message persistence table (econ_bank_messages)
# - No secondary_type, no ECON_ADMIN_ROLE_IDS, no duplicate purchase commands, no global slash sync (guild-only)
# - All slash commands defer immediately

from __future__ import annotations
import unicodedata

import os
import json
import re
import asyncio
from io import BytesIO
from pathlib import Path
from dataclasses import dataclass
from datetime import date, datetime, timezone
from typing import Dict, List, Optional, Tuple, Any

import discord
from discord import app_commands
from discord.ext import tasks

try:
    from zoneinfo import ZoneInfo  # py>=3.9
except Exception:
    ZoneInfo = None  # type: ignore

try:
    import asyncpg
except Exception as e:
    raise RuntimeError("asyncpg is required for EconBot") from e


try:
    import openpyxl  # type: ignore
except Exception:
    openpyxl = None  # type: ignore


APP_VERSION = "EconBot_v143"

# Canon kingdoms (authoritative list for tax dropdowns & treasury seeding)
CANON_KINGDOMS: list[str] = ["Sethrathiel", "Velarith", "Lyvik", "Baelon", "Avalea"]
DEFAULT_KINGDOM_TAX_BP = 1000  # 10%
CHICAGO_TZ = ZoneInfo("America/Chicago") if ZoneInfo else timezone.utc

# Bundled images (do NOT use expiring Discord CDN URLs).
# These are relative to the repo root beside main.py, typically under ./assets/.
BALANCE_CARD_IMAGE_PATH = "assets/great_anus.jpg"
BALANCE_CARD_IMAGE_FILENAME = "balance_card.jpg"
WEEKEND_INCOME_IMAGE_PATH = "assets/weekend_income.jpg"
WEEKEND_INCOME_IMAGE_FILENAME = "weekend_income.jpg"

# Daily income reminder (hard-coded by request; no Railway env vars required).
DAILY_REMINDER_CHANNEL_ID = 1324994929176612936
DAILY_REMINDER_ROLE_ID = 1476435497776840724
DAILY_REMINDER_HOUR = 12
DAILY_REMINDER_MINUTE = 0
DAILY_REMINDER_MESSAGE = f"<@&{DAILY_REMINDER_ROLE_ID}> - Don’t forget to claim your daily income (and pay your taxes) using the /income command!"

# Resolve bundled image paths relative to this file (Railway working directory can vary).
_BASE_DIR = Path(__file__).resolve().parent


def _find_bundled_image(config_path: str, fallback_names: list[str]) -> Optional[Path]:
    """Return an existing bundled image path, or None if not found."""
    candidates: list[Path] = []
    p = Path(config_path)
    candidates.append(p if p.is_absolute() else (_BASE_DIR / p))
    assets_dir = _BASE_DIR / "assets"
    for name in fallback_names:
        candidates.append(assets_dir / name)
    for c in candidates:
        try:
            if c.exists() and c.is_file():
                return c
        except Exception:
            continue
    return None


def _image_embed(filename: str) -> discord.Embed:
    e = discord.Embed()
    e.set_image(url=f"attachment://{filename}")
    return e


async def send_ephemeral_with_bundled_image(
    interaction: discord.Interaction,
    content: str,
    *,
    config_path: str,
    filename: str,
    fallback_names: list[str],
    label: str,
):
    p = _find_bundled_image(config_path, fallback_names)
    if not p:
        print(f"[warn] {label} image missing. Tried {config_path!r} relative to {_BASE_DIR}.")
        assets_dir = _BASE_DIR / "assets"
        if assets_dir.exists() and assets_dir.is_dir():
            try:
                print(f"[debug] assets/ contents: {[x.name for x in assets_dir.iterdir() if x.is_file()]}")
            except Exception:
                pass
        await interaction.followup.send(content, ephemeral=True, allowed_mentions=discord.AllowedMentions.none())
        return
    try:
        f = discord.File(str(p), filename=filename)
        await interaction.followup.send(
            content,
            ephemeral=True,
            embed=_image_embed(filename),
            file=f,
            allowed_mentions=discord.AllowedMentions.none(),
        )
    except FileNotFoundError:
        print(f"[warn] {label} image missing at {str(p)!r}; sending without image")
        await interaction.followup.send(content, ephemeral=True, allowed_mentions=discord.AllowedMentions.none())


async def send_ephemeral_with_parchment(interaction: discord.Interaction, content: str):
    await send_ephemeral_with_bundled_image(
        interaction,
        content,
        config_path=BALANCE_CARD_IMAGE_PATH,
        filename=BALANCE_CARD_IMAGE_FILENAME,
        fallback_names=[
            "great_anus.jpg",
            "great_anus.jpg.jpg",
            "Great Anus.jpg",
            "Great Anus (1).jpg",
            "Great%20Anus.jpg",
            "great_anus.jpeg",
            "great_anus.png",
        ],
        label="Balance/Income",
    )


async def send_ephemeral_with_weekend_income_image(interaction: discord.Interaction, content: str):
    await send_ephemeral_with_bundled_image(
        interaction,
        content,
        config_path=WEEKEND_INCOME_IMAGE_PATH,
        filename=WEEKEND_INCOME_IMAGE_FILENAME,
        fallback_names=[
            "weekend_income.jpg",
            "weekend_income.jpeg",
            "BA4EB710-C8AB-4B8F-AFDC-EDBAFDEE13F6.jpeg",
        ],
        label="Weekend income",
    )


# -------------------------

# --- Internal throttles (prevent rate-limit cascades) ---
_STAFF_MEMBER_FETCH_LOCK = asyncio.Lock()

# Env helpers (NO inventions)
# -------------------------

def _get(name: str, default: Optional[str] = None) -> Optional[str]:
    v = os.getenv(name)
    if v is None:
        return default
    v = v.strip()
    return v if v != "" else default


def _int(name: str, default: Optional[int] = None) -> Optional[int]:
    raw = _get(name)
    if raw is None:
        return default
    digits = re.sub(r"[^0-9]", "", raw)
    if digits == "":
        return default
    try:
        return int(digits)
    except Exception:
        return default


def _int_list(name: str) -> List[int]:
    raw = _get(name, "")
    parts = [p.strip() for p in str(raw).replace("\n", ",").replace(";", ",").split(",") if p.strip()]
    out: List[int] = []
    for p in parts:
        digits = re.sub(r"[^0-9]", "", p)
        if digits:
            try:
                out.append(int(digits))
            except Exception:
                pass
    # de-dupe stable
    return sorted(list(dict.fromkeys(out)))


def _tier_rank(tier: str) -> Optional[int]:
    """Extract numeric rank from tier labels like '(3) Small Tavern'."""
    if tier is None:
        return None
    s = str(tier).strip()
    m = re.match(r"^\(\s*(\d+)\s*\)", s)
    if not m:
        m = re.match(r"^(\d+)", s)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


async def cumulative_cost_to_tier(asset_type: str, target_tier: str) -> Optional[int]:
    """Cumulative cost: sum of all tier costs up to and including target tier for an asset type.

    Requires tiers to be ordered by a numeric rank embedded in the tier label, e.g. '(1) ...', '(2) ...'.
    Falls back to the target tier's own cost if ranks are not parseable.
    """
    target_rank = _tier_rank(target_tier)

    rows = await db_fetch(
        '''
        SELECT tier, cost_val
        FROM econ_asset_definitions
        WHERE asset_type=$1;
        ''',
        asset_type,
    )
    if not rows:
        return None

    # Normalize tier strings once
    norm_rows = []
    for r in rows:
        t = str(r["tier"]).strip()
        try:
            c = int(r["cost_val"])
        except Exception:
            continue
        norm_rows.append((t, c, _tier_rank(t)))

    # If the selected tier has a rank, sum all rows with rank <= target_rank
    if target_rank is not None:
        total = 0
        found = False
        for t, c, rk in norm_rows:
            if rk is None:
                continue
            if rk <= target_rank:
                total += c
                found = True
        if found:
            return total

    # Fallback: just use the selected tier's own cost
    target_norm = str(target_tier).strip()
    for t, c, rk in norm_rows:
        if t == target_norm:
            return c

    return None



def _plural(unit: str, qty: int) -> str:
    """Pluralization for currency units.

    Canon (per user):
    - Elsh plural is Elsh
    - Oril plural is Orils
    - Arce plural is Arces
    - Cinth plural is Cinths
    - Novir shown as Novir for both (kept as-is)
    """
    if qty == 1:
        return unit
    if unit == "Elsh":
        return "Elsh"
    if unit == "Oril":
        return "Orils"
    if unit == "Arce":
        return "Arces"
    if unit == "Cinth":
        return "Cinths"
    if unit == "Novir":
        return "Novir"
    return unit + "s"


def format_balance(total_val: int) -> str:
    """Format as: '1 Novir, 2 Orils, 3 Elsh, 4 Arces, 5 Cinths (12,345 Val)'.

    Always shows all denominations (including zeros) and includes total in Val.
    """
    try:
        total = int(total_val)
    except Exception:
        total = 0

    sign = "-" if total < 0 else ""
    n = abs(total)

    novir, n = divmod(n, 10000)
    oril, n = divmod(n, 1000)
    elsh, n = divmod(n, 100)
    arce, cinth = divmod(n, 10)

    parts = [
        f"{novir} {_plural('Novir', novir)}",
        f"{oril} {_plural('Oril', oril)}",
        f"{elsh} {_plural('Elsh', elsh)}",
        f"{arce} {_plural('Arce', arce)}",
        f"{cinth} {_plural('Cinth', cinth)}",
    ]
    return f"{sign}{', '.join(parts)} ({total:,} Val)"


def sanitize_plain_text(s: str) -> str:
    """Sanitize user/DB-provided text so Discord markdown can't be broken by backticks/codeblocks.
    We do NOT escape '*' or '_' globally (you rely on those for formatting).
    We only neutralize characters that can force code blocks or inline code, and strip zero-width chars.
    """
    if s is None:
        return ""
    s = str(s)
    # Normalize and remove zero-width/invisible formatting characters
    s = unicodedata.normalize("NFKC", s)
    s = "".join(ch for ch in s if ch not in {"\u200b", "\u200c", "\u200d", "\ufeff"})
    # Neutralize backticks which can create inline code or code blocks
    s = s.replace("```", "''").replace("`", "'")
    # Prevent accidental mentions even though we also set AllowedMentions.none()
    s = s.replace("@", "@\u200b")
    # Hard-trim
    return s.strip()


def normalize_dashboard_markdown(s: str) -> str:
    """Ensure Discord renders markdown by preventing accidental code blocks.
    - Removes leading BOM/zero-width chars
    - Strips any accidental triple-backtick fences
    - Removes 4-space indentation that would create preformatted blocks
    - Strips trailing whitespace
    """
    if s is None:
        return ""
    # Normalize newlines
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    # Remove any accidental fenced code blocks
    s = s.replace("```", "''")
    # Clean lines
    out_lines = []
    for line in s.split("\n"):
        # Strip zero-width / BOM at start of line
        line = line.lstrip("\ufeff\u200b\u200c\u200d")
        # Prevent 4-space indentation from triggering preformatted/code blocks
        if line.startswith("    "):
            line = line.lstrip()
        out_lines.append(line.rstrip())
    # Ensure message doesn't start with whitespace
    return "\n".join(out_lines).strip("\n")



def format_amount(val: int) -> str:
    """Single-amount formatter for income lines and per-asset deltas.

    - If exactly 1 Arce (10 Val), show '1 Arce'.
    - Otherwise show '{val:,} Val'.
    """
    try:
        v = int(val)
    except Exception:
        v = 0
    if v == 10:
        return "1 Arce"
    return f"{v:,} Val"


def format_currency(total_val: int) -> str:
    """Backward-compatible compact formatter used across non-card views.

    Keeps existing parts of the bot (treasuries, leaderboards, logs) stable.
    """
    try:
        total = int(total_val)
    except Exception:
        total = 0

    sign = "-" if total < 0 else ""
    n = abs(total)

    denominations = [
        (10000, "Novir"),
        (1000, "Oril"),
        (100, "Elsh"),
        (10, "Arce"),
        (1, "Cinth"),
    ]

    parts: List[str] = []
    for value, unit in denominations:
        if n <= 0:
            break
        qty, n = divmod(n, value)
        if qty:
            parts.append(f"{qty} {_plural(unit, int(qty))}")

    if not parts:
        parts = ["0 Cinths"]

    compact = " • ".join(parts)
    return f"{sign}{compact} ({total:,} Val)"

# Base daily income granted on /income claim (in Copper Cinth units)
BASE_DAILY_INCOME = 10




def current_base_daily_income(now: Optional[datetime] = None) -> int:
    """Return base daily income, doubled on Fri/Sat/Sun (America/Chicago)."""
    dt = now or datetime.now(CHICAGO_TZ)
    return int(BASE_DAILY_INCOME) * (2 if dt.weekday() in (4, 5, 6) else 1)



async def tier_cost_for(asset_type: str, tier: str) -> Optional[int]:
    """Return cost_val for the exact tier (NOT cumulative)."""
    row = await db_fetchrow(
        '''
        SELECT cost_val
        FROM econ_asset_definitions
        WHERE asset_type=$1 AND tier=$2
        LIMIT 1;
        ''',
        asset_type,
        tier,
    )
    if not row:
        return None
    try:
        return int(row["cost_val"])
    except Exception:
        return None


async def incremental_cost_between_tiers(asset_type: str, current_tier: str, target_tier: str) -> Optional[int]:
    """Upgrade cost from current_tier to target_tier for an asset_type.

    Rule: sum costs of tiers strictly above current tier through the target tier.
    Fallback: cumulative(target) - cumulative(current) if rank-based sum cannot be computed.
    """
    cur_rank = _tier_rank(current_tier)
    tgt_rank = _tier_rank(target_tier)
    if tgt_rank is None:
        return None

    rows = await db_fetch(
        '''
        SELECT tier, cost_val
        FROM econ_asset_definitions
        WHERE asset_type=$1;
        ''',
        asset_type,
    )
    if not rows:
        return None

    if cur_rank is not None:
        total = 0
        any_found = False
        for r in rows:
            t = str(r["tier"])
            tr = _tier_rank(t)
            if tr is None:
                continue
            if tr > cur_rank and tr <= tgt_rank:
                any_found = True
                total += int(r["cost_val"])
        if any_found:
            return int(total)

    cum_t = await cumulative_cost_to_tier(asset_type, target_tier)
    cum_c = await cumulative_cost_to_tier(asset_type, current_tier)
    if cum_t is None or cum_c is None:
        return None
    return int(cum_t - cum_c)

async def get_assets_for_character(character_name: str) -> List[Dict[str, Any]]:
    """Return full asset rows for a character."""
    rows = await db_fetch(
        '''
        SELECT asset_type, tier, asset_name, COALESCE(kingdom, '') AS kingdom
        FROM econ_assets
        WHERE guild_id=$1 AND character_name=$2
        ''',
        DATA_GUILD_ID,
        character_name,
    )
    # sort by asset_type, then tier rank, then name
    def _key(r: Dict[str, Any]) -> Tuple[str, int, str]:
        t = str(r.get("tier", ""))
        rk = _tier_rank(t)
        if rk is None:
            rk = 9999
        return (str(r.get("asset_type", "")), rk, str(r.get("asset_name", "")))
    return sorted(rows, key=_key)


async def get_assets_with_income_for_character(character_name: str) -> List[Dict[str, Any]]:
    """Return assets for a character including per-tier income value."""
    rows = await db_fetch(
        '''
        SELECT a.asset_type, a.tier, a.asset_name, COALESCE(a.kingdom, '') AS kingdom,
               COALESCE(a.noble_title_family, '') AS noble_title_family,
               COALESCE(a.noble_title_option, '') AS noble_title_option,
               COALESCE(a.noble_realm_name, '') AS noble_realm_name,
               d.add_income_val
        FROM econ_assets a
        JOIN econ_asset_definitions d
          ON d.asset_type=a.asset_type AND d.tier=a.tier
        WHERE a.guild_id=$1 AND a.character_name=$2
        ''',
        DATA_GUILD_ID,
        character_name,
    )

    def _key(r: Dict[str, Any]) -> Tuple[int, str, str]:
        rk = _tier_rank(str(r.get("tier", "")))
        if rk is None:
            rk = 9999
        return (rk, str(r.get("asset_type", "")), str(r.get("asset_name", "")))

    return sorted(rows, key=_key)


async def _display_name_from_cache(guild: discord.Guild, user_id: int) -> str:
    """Plain-text display name (server nickname if set). No pings, no fetch."""
    m = guild.get_member(int(user_id))
    return m.display_name if m else f"User {user_id}"


def _tier_label(tier: str) -> str:
    rk = _tier_rank(str(tier))
    if rk is None:
        return str(tier)
    return f"T{rk}"



def noble_title_family_requires_realm(family: str) -> bool:
    meta = NOBLE_TITLE_FAMILIES.get(str(family or "").strip())
    if not meta:
        return True
    return bool(meta.get("requires_realm", True))


def noble_title_option_requires_realm(option: str, family: str = "") -> bool:
    fam = str(family or "").strip()
    if fam and fam in NOBLE_TITLE_FAMILIES:
        return noble_title_family_requires_realm(fam)
    return str(option or "").strip() not in SOVEREIGN_TITLE_OPTIONS


def noble_title_display_from_parts(option: str, realm: str, kingdom: str, family: str = "") -> str:
    option = sanitize_plain_text(str(option or "")).strip()
    realm = sanitize_plain_text(str(realm or "")).strip()
    kingdom = sanitize_plain_text(str(kingdom or "")).strip()
    requires_realm = noble_title_option_requires_realm(option, family)
    if requires_realm:
        if realm and kingdom:
            return f"{option} of {realm} | Kingdom of {kingdom}"
        if realm:
            return f"{option} of {realm}"
        if kingdom:
            return f"{option} | Kingdom of {kingdom}"
        return option or "(Untitled)"
    if kingdom:
        return f"{option} of {kingdom}"
    return option or "(Untitled)"


def noble_title_display_from_row(row: Dict[str, Any]) -> str:
    option = str(row.get("noble_title_option") or row.get("asset_name") or "")
    realm = str(row.get("noble_realm_name") or "")
    kingdom = str(row.get("kingdom") or "")
    family = str(row.get("noble_title_family") or TIER_TO_NOBLE_TITLE_FAMILY.get(str(row.get("tier") or "")) or "")
    return noble_title_display_from_parts(option, realm, kingdom, family)


def noble_title_identifier_from_row(row: Dict[str, Any]) -> str:
    option = sanitize_plain_text(str(row.get("noble_title_option") or row.get("asset_name") or "")).strip()
    realm = sanitize_plain_text(str(row.get("noble_realm_name") or "")).strip()
    kingdom = sanitize_plain_text(str(row.get("kingdom") or "")).strip()
    return f"{option}|||{realm}|||{kingdom}"[:100]


def parse_noble_title_identifier(identifier: str) -> Tuple[str, str, str]:
    raw = str(identifier or "")
    parts = raw.split("|||", 2)
    while len(parts) < 3:
        parts.append("")
    return tuple(sanitize_plain_text(p).strip() for p in parts[:3])


def noble_title_rank_for_family(family: str) -> int:
    meta = NOBLE_TITLE_FAMILIES.get(str(family or "").strip())
    if not meta:
        return 0
    return int(meta["rank"])


def noble_title_cost_for_family(family: str) -> int:
    meta = NOBLE_TITLE_FAMILIES.get(str(family or "").strip())
    if not meta:
        return 0
    return int(meta["cost"])


async def get_noble_titles_for_character(character_name: str) -> List[Dict[str, Any]]:
    rows = await db_fetch(
        """
        SELECT asset_type, tier, asset_name, COALESCE(kingdom, '') AS kingdom,
               COALESCE(noble_title_family, '') AS noble_title_family,
               COALESCE(noble_title_option, '') AS noble_title_option,
               COALESCE(noble_realm_name, '') AS noble_realm_name
        FROM econ_assets
        WHERE guild_id=$1 AND character_name=$2 AND asset_type='Noble Title'
        ORDER BY COALESCE(noble_realm_name, ''), tier, asset_name;
        """,
        DATA_GUILD_ID,
        character_name,
    )
    return rows


async def render_character_section(character_name: str) -> List[str]:
    """Render the character portion of a card (no nickname header, no border)."""
    kingdom = await get_character_kingdom(character_name)
    kingdom = sanitize_plain_text(str(kingdom or "")).strip() or "(No Kingdom)"

    bal = await get_balance(character_name)
    assets = await get_assets_with_income_for_character(character_name)
    asset_income_sum = sum(int(a.get("add_income_val") or 0) for a in assets)
    title_assets = [a for a in assets if is_noble_title_asset_type(str(a.get("asset_type", "")))]
    regular_assets = [a for a in assets if not is_noble_title_asset_type(str(a.get("asset_type", "")))]

    out: List[str] = []
    safe_cname = sanitize_plain_text(character_name)
    out.append(f"**{safe_cname}** - {kingdom}")
    out.append(f"💰 **Balance:** {format_balance(bal)}")
    out.append(f"📈 **Income:** {format_amount(current_base_daily_income())} | **Income from Assets:** {asset_income_sum:,} Val")

    if title_assets:
        out.append("🏛 **Noble Titles**")
        for a in title_assets:
            out.append(f"- {noble_title_display_from_row(a)}")

    out.append("🧾 **Assets**")

    if not regular_assets:
        out.append("- (none)")
        return out

    for a in regular_assets:
        tier = _tier_label(str(a.get("tier", "")))
        aname = sanitize_plain_text(str(a.get("asset_name", ""))).strip() or sanitize_plain_text(str(a.get("asset_type", ""))).strip()
        akingdom = sanitize_plain_text(str(a.get("kingdom", ""))).strip() or "(No Kingdom)"
        add = int(a.get("add_income_val") or 0)
        out.append(f"- {tier} - {aname} - {akingdom} - +{add:,} Val")
    return out


async def render_user_card_block(
    guild: discord.Guild,
    owner_id: int,
    character_names: List[str],
) -> List[str]:
    """Render a full card block matching the user's required Discord formatting."""
    owner_display = sanitize_plain_text(await _display_name_from_cache(guild, owner_id))

    out: List[str] = []
    out.append("___________________________________________________________________")
    out.append(f"***{owner_display}***")

    for idx, cname in enumerate(character_names):
        if idx > 0:
            out.append("")
        out.extend(await render_character_section(cname))

    out.append("___________________________________________________________________")
    return out


async def render_leaderboard_lines(
    guild: discord.Guild,
    rows: List[Tuple[str, int, int, int]],
) -> List[str]:
    """rows: [(character_name, owner_id, balance, income)]"""
    # Top balances
    top_bal = sorted(rows, key=lambda r: int(r[2]), reverse=True)[:5]
    top_inc = sorted(rows, key=lambda r: int(r[3]), reverse=True)[:5]

    # resolve owner display names (no pings)
    cache: Dict[int, str] = {}
    async def dn(uid: int) -> str:
        if uid in cache:
            return cache[uid]
        m = guild.get_member(uid)
        # Do not fetch members (avoids rate limits); rely on cache only.
        cache[uid] = m.display_name if m else f"User {uid}"
        return cache[uid]

    out: List[str] = []
    out.append("🏆 **Leaderboards**")
    out.append("━━━━━━━━━━━━━━━━━━")
    out.append("**Top Balances**")
    if not top_bal:
        out.append("- (none)")
    else:
        for i, (c, uid, bal, inc) in enumerate(top_bal, start=1):
            out.append(f"{i}. **{sanitize_plain_text(c)}** — *{sanitize_plain_text(await dn(uid))}* — {format_currency(bal)}")
    out.append("")
    out.append("**Top Daily Income**")
    if not top_inc:
        out.append("- (none)")
    else:
        for i, (c, uid, bal, inc) in enumerate(top_inc, start=1):
            out.append(f"{i}. **{sanitize_plain_text(c)}** — *{sanitize_plain_text(await dn(uid))}* — {format_currency(inc)}")
    out.append("")
    return out


# Approved env vars (per continuity doc)
DISCORD_TOKEN = _get("DISCORD_TOKEN")
DATABASE_URL = _get("DATABASE_URL")
GUILD_ID = _int("GUILD_ID")
LEGACY_SOURCE_GUILD_ID = _int("LEGACY_SOURCE_GUILD_ID", GUILD_ID)
BANK_CHANNEL_ID = _int("BANK_CHANNEL_ID")
ECON_LOG_CHANNEL_ID = _int("ECON_LOG_CHANNEL_ID")
STAFF_ROLE_IDS = set(_int_list("STAFF_ROLE_IDS"))
STAFF_ROLE_IDS_DEFAULT = {1473523681132019824, 1473523738891784232}  # fallback if env missing
if not STAFF_ROLE_IDS:
    STAFF_ROLE_IDS = set(STAFF_ROLE_IDS_DEFAULT)

BANK_REFRESH_ROLE_IDS = set(_int_list("BANK_REFRESH_ROLE_IDS"))
# BANK_MESSAGE_IDS is kept for backward compatibility, but v72 persists message IDs in Postgres (approved)
BANK_MESSAGE_IDS = _int_list("BANK_MESSAGE_IDS")

if not DISCORD_TOKEN:
    raise RuntimeError("Missing DISCORD_TOKEN")
if not DATABASE_URL:
    raise RuntimeError("Missing DATABASE_URL")
if not GUILD_ID:
    raise RuntimeError("Missing/invalid GUILD_ID")
if not LEGACY_SOURCE_GUILD_ID:
    raise RuntimeError("Missing/invalid LEGACY_SOURCE_GUILD_ID (or GUILD_ID)")

# All DB reads/writes (balances/assets/income/characters) use the legacy guild partition, per your confirmation.
DATA_GUILD_ID = int(LEGACY_SOURCE_GUILD_ID)



# -------------------------
# Asset definitions seed (authoritative from NEW Asset Table.xlsx)
# -------------------------

NOBLE_TITLE_FAMILIES: Dict[str, Dict[str, Any]] = {
    "Baron/ess": {"tier": "(1) Baron/ess", "rank": 1, "cost": 4000, "options": ["Baron", "Baroness"], "requires_realm": True},
    "Viscount/ess": {"tier": "(2) Viscount/ess", "rank": 2, "cost": 6000, "options": ["Viscount", "Viscountess"], "requires_realm": True},
    "Earl/Countess": {"tier": "(3) Earl/Countess", "rank": 3, "cost": 9000, "options": ["Earl", "Countess"], "requires_realm": True},
    "Marquess/Marchioness": {"tier": "(4) Marquess/Marchioness", "rank": 4, "cost": 13000, "options": ["Marquess", "Marchioness"], "requires_realm": True},
    "Duke/Duchess": {"tier": "(5) Duke/Duchess", "rank": 5, "cost": 20000, "options": ["Duke", "Duchess"], "requires_realm": True},
    "Regent": {"tier": "(6) Regent", "rank": 6, "cost": 0, "options": ["Regent"], "requires_realm": False},
    "King/Queen": {"tier": "(7) King/Queen", "rank": 7, "cost": 0, "options": ["King", "Queen"], "requires_realm": False},
    "Sovereign": {"tier": "(8) Sovereign", "rank": 8, "cost": 0, "options": ["Sovereign"], "requires_realm": False},
}

NOBLE_TITLE_DEFINITIONS: List[Tuple[str, str, int, int]] = [
    ("Noble Title", str(meta["tier"]), int(meta["cost"]), 0)
    for _family, meta in NOBLE_TITLE_FAMILIES.items()
]

NOBLE_TITLE_LOOKUP: Dict[str, str] = {
    family: str(meta["tier"]) for family, meta in NOBLE_TITLE_FAMILIES.items()
}
NOBLE_TITLE_OPTION_TO_FAMILY: Dict[str, str] = {
    opt: family
    for family, meta in NOBLE_TITLE_FAMILIES.items()
    for opt in meta["options"]
}
TIER_TO_NOBLE_TITLE_FAMILY: Dict[str, str] = {
    str(meta["tier"]): family for family, meta in NOBLE_TITLE_FAMILIES.items()
}
SPECIAL_NO_COST_TITLE_FAMILIES = {"Regent", "King/Queen", "Sovereign"}
SOVEREIGN_TITLE_OPTIONS = {"Regent", "King", "Queen", "Sovereign"}


def is_noble_title_asset_type(asset_type: str) -> bool:
    return str(asset_type or "").strip().lower() == "noble title"


ASSET_DEFINITIONS_SEED: List[Tuple[str, str, int, int]] = [
    ("Guild Trade Workshop", "(1) Guild Apprentice", 300, 50),
    ("Guild Trade Workshop", "(2) Guild Journeyman", 600, 100),
    ("Guild Trade Workshop", "(3) Leased Workshop", 1200, 150),
    ("Guild Trade Workshop", "(4) Small Workshop", 2000, 200),
    ("Guild Trade Workshop", "(5) Large Workshop", 3000, 250),
    ("Market Stall", "(1) Consignment Arrangement", 300, 50),
    ("Market Stall", "(2) Small Alley Stand", 600, 100),
    ("Market Stall", "(3) Market Stall", 1200, 150),
    ("Market Stall", "(4) Small Shop", 2000, 200),
    ("Market Stall", "(5) Large Shop", 3000, 250),
    ("Farm/Ranch", "(1) Subsistence Surplus", 300, 50),
    ("Farm/Ranch", "(2) Leased Fields", 600, 100),
    ("Farm/Ranch", "(3) Owned Acre", 1200, 150),
    ("Farm/Ranch", "(4) Small Fields and Barn", 2000, 200),
    ("Farm/Ranch", "(5) Large Fields and Barn", 3000, 250),
    ("Tavern/Inn", "(1) One-Room Flophouse", 300, 50),
    ("Tavern/Inn", "(2) Leased Establishment", 600, 100),
    ("Tavern/Inn", "(3) Small Tavern", 1200, 150),
    ("Tavern/Inn", "(4) Large Tavern", 2000, 200),
    ("Tavern/Inn", "(5) Large Tavern and Inn", 3000, 250),
    ("Warehouse/Trade House", "(1) Small Storage Shed", 300, 50),
    ("Warehouse/Trade House", "(2) Large Storage Shed", 600, 100),
    ("Warehouse/Trade House", "(3) Small Trading Post", 1200, 150),
    ("Warehouse/Trade House", "(4) Large Trading Post", 2000, 200),
    ("Warehouse/Trade House", "(5) Large Warehouse and Trading Post", 3000, 250),
    ("House", "(1) Shack", 600, 0),
    ("House", "(2) Hut", 1200, 0),
    ("House", "(3) House", 2000, 0),
    ("House", "(4) Lodge", 3000, 0),
    ("House", "(5) Mansion", 5000, 0),
    ("Village", "(1) Chartered Assembly", 1200, 100),
    ("Village", "(2) Hamlet", 2400, 200),
    ("Village", "(3) Village", 4800, 300),
    ("Village", "(4) Town", 9600, 400),
    ("Village", "(5) Small City", 15000, 500),
    ("Weapons", "(1) Hit +1 / Dmg +1d4", 300, 0),
    ("Weapons", "(2) Hit +1 / Dmg +1d6", 600, 0),
    ("Weapons", "(3) Hit +2 / Dmg +1d8", 1200, 0),
    ("Weapons", "(4) Hit +2 / Dmg +1d10", 2400, 0),
    ("Weapons", "(5) Hit +2 / Dmg +1d12", 4800, 0),
    ("Armor", "(1) AC +1", 300, 0),
    ("Armor", "(2) AC +2", 600, 0),
    ("Armor", "(3) AC +2 / Adv Magic Atk", 1200, 0),
    ("Armor", "(4) AC +2 / Adv Magic and Melee Atk", 2400, 0),
    ("Armor", "(5) AC +3 / Adv Magic and Melee Atk", 4800, 0),
] + NOBLE_TITLE_DEFINITIONS

# -------------------------
# Discord client setup
# -------------------------

intents = discord.Intents.default()
intents.members = True  # required for role detection
client = discord.Client(intents=intents)
tree = app_commands.CommandTree(client)




def _cap_message(text: str, limit: int = 1900) -> str:
    """Cap a Discord message to avoid 2000-char hard limit. Adds ellipsis when trimmed."""
    if text is None:
        return ""
    if len(text) <= limit:
        return text
    trimmed = text[: max(0, limit - 20)]
    nl = trimmed.rfind("\n")
    if nl > 200:
        trimmed = trimmed[:nl]
    return trimmed.rstrip() + "\n… _(truncated)_"

# -------------------------
# DB
# -------------------------

_POOL: Optional["asyncpg.Pool"] = None


async def db_pool() -> "asyncpg.Pool":
    global _POOL
    if _POOL is None:
        _POOL = await asyncpg.create_pool(DATABASE_URL, min_size=1, max_size=5, command_timeout=30)
    return _POOL


async def db_exec(sql: str, *args) -> str:
    pool = await db_pool()
    async with pool.acquire() as conn:
        return await conn.execute(sql, *args)


async def db_fetch(sql: str, *args) -> List[asyncpg.Record]:
    pool = await db_pool()
    async with pool.acquire() as conn:
        return await conn.fetch(sql, *args)


async def db_fetchrow(sql: str, *args) -> Optional[asyncpg.Record]:
    pool = await db_pool()
    async with pool.acquire() as conn:
        return await conn.fetchrow(sql, *args)


# -------------------------
# Migrations (minimal, explicit)
# -------------------------
# ONLY new table approved by you:
# econ_bank_messages (to persist bank dashboard message IDs instead of env var spam)


async def ensure_schema() -> None:
    # Core tables expected to already exist (v38 doctrine). We do NOT create them here.
    # We *do* create the approved bank-message persistence table.
    await db_exec(
        """
        CREATE TABLE IF NOT EXISTS econ_bank_messages (
            guild_id BIGINT NOT NULL,
            idx INTEGER NOT NULL,
            message_id BIGINT NOT NULL,
            PRIMARY KEY (guild_id, idx)
        );
        """
    )

    await db_exec(
        """
        CREATE TABLE IF NOT EXISTS econ_daily_reminder_runs (
            guild_id BIGINT NOT NULL,
            reminder_date DATE NOT NULL,
            PRIMARY KEY (guild_id, reminder_date)
        );
        """
    )

    # Kingdom taxation support (explicitly approved):
    # - characters.kingdom (required for income claims; populated by upstream character-creation bot)
    # - econ_assets.kingdom (optional override; if NULL, inherits character home kingdom)
    # - econ_kingdoms treasury + tax rates (basis points)
    try:
        await db_exec("ALTER TABLE characters ADD COLUMN IF NOT EXISTS kingdom TEXT;")
    except Exception as e:
        print(f"[warn] Could not add characters.kingdom (will require manual migration): {e}")

    try:
        await db_exec("ALTER TABLE econ_assets ADD COLUMN IF NOT EXISTS kingdom TEXT;")
    except Exception as e:
        print(f"[warn] Could not add econ_assets.kingdom (will require manual migration): {e}")

    # Noble title metadata (additive only; required by v129 title rendering/commands)
    try:
        await db_exec("ALTER TABLE econ_assets ADD COLUMN IF NOT EXISTS noble_title_family TEXT;")
        await db_exec("ALTER TABLE econ_assets ADD COLUMN IF NOT EXISTS noble_title_option TEXT;")
        await db_exec("ALTER TABLE econ_assets ADD COLUMN IF NOT EXISTS noble_realm_name TEXT;")
    except Exception as e:
        print(f"[warn] Could not add econ_assets noble-title columns (will require manual migration): {e}")

    await db_exec(
        """
        CREATE TABLE IF NOT EXISTS econ_kingdoms (
            guild_id BIGINT NOT NULL,
            kingdom TEXT NOT NULL,
            tax_rate_bp INT NOT NULL DEFAULT 0,
            treasury BIGINT NOT NULL DEFAULT 0,
            PRIMARY KEY (guild_id, kingdom)
        );
        """
    )

    # Seed canonical kingdoms with baseline 10% tax (do not override non-zero rates)
    try:
        for _k in CANON_KINGDOMS:
            await db_exec(
                """
                INSERT INTO econ_kingdoms (guild_id, kingdom, tax_rate_bp, treasury)
                VALUES ($1, $2, $3, 0)
                ON CONFLICT (guild_id, kingdom) DO UPDATE
                SET tax_rate_bp = CASE WHEN econ_kingdoms.tax_rate_bp = 0 THEN EXCLUDED.tax_rate_bp ELSE econ_kingdoms.tax_rate_bp END;
                """,
                DATA_GUILD_ID,
                _k,
                DEFAULT_KINGDOM_TAX_BP,
            )
    except Exception as e:
        print(f"[warn] Could not seed econ_kingdoms baseline rates: {e}")



    # Adjust econ_assets unique constraint so the same asset_name can be reused across different asset_type/tier.
    # Desired uniqueness: (guild_id, user_id, character_name, asset_type, tier, asset_name)
    try:
        rows = await db_fetch(
            '''
            SELECT conname, pg_get_constraintdef(c.oid) AS def
            FROM pg_constraint c
            JOIN pg_class t ON t.oid = c.conrelid
            WHERE t.relname = 'econ_assets' AND c.contype = 'u';
            '''
        )
        for r in rows:
            conname = str(r["conname"])
            cdef = str(r["def"]).replace('"', "")
            if "UNIQUE" in cdef and "(guild_id, user_id, character_name, asset_name)" in cdef:
                await db_exec(f'ALTER TABLE econ_assets DROP CONSTRAINT IF EXISTS "{conname}";')
                print(f"[test] Dropped old econ_assets unique constraint: {conname}")

        exists = await db_fetchrow(
            '''
            SELECT 1
            FROM pg_constraint c
            JOIN pg_class t ON t.oid = c.conrelid
            WHERE t.relname='econ_assets' AND c.contype='u' AND c.conname='econ_assets_unique_name_per_type_tier'
            LIMIT 1;
            '''
        )
        if not exists:
            await db_exec(
                '''
                ALTER TABLE econ_assets
                ADD CONSTRAINT econ_assets_unique_name_per_type_tier
                UNIQUE (guild_id, user_id, character_name, asset_type, tier, asset_name);
                '''
            )
            print("[test] Added unique constraint econ_assets_unique_name_per_type_tier")
    except Exception as e:
        print(f"[warn] econ_assets unique-constraint adjustment skipped/failed: {e}")


async def _get_member(interaction: discord.Interaction) -> Optional[discord.Member]:
    """Resolve the invoking member for staff permission checks.

    Priority order:
    1) Use interaction.user if it's already a discord.Member (includes roles).
    2) Use guild cache lookup (no HTTP).
    3) As a last resort, fetch the member over HTTP *once* (serialized) to avoid false denials.
       This is intentionally gated behind a lock to reduce 429s.
    """
    if isinstance(interaction.user, discord.Member):
        return interaction.user

    if interaction.guild is None:
        return None

    try:
        cached = interaction.guild.get_member(interaction.user.id)
        if cached is not None:
            return cached
    except Exception:
        pass

    # Last resort: single serialized HTTP fetch (helps when members intent/caching is insufficient).
    try:
        async with _STAFF_MEMBER_FETCH_LOCK:
            return await interaction.guild.fetch_member(interaction.user.id)
    except Exception:
        return None

    # As a last resort, do NOT fetch over HTTP here; return None and staff gate will explain.
    return None


async def is_staff(interaction: discord.Interaction) -> Tuple[bool, Dict[str, Any]]:
    """Return (is_staff, debug_dict). Uses role IDs and falls back to guild permissions."""
    member = await _get_member(interaction)

    role_ids: List[int] = []
    admin = False
    manage_guild = False
    manage_messages = False

    if member is not None:
        try:
            role_ids = [int(r.id) for r in getattr(member, "roles", [])]
        except Exception:
            role_ids = []
        try:
            perms = getattr(member, "guild_permissions", None)
            if perms is not None:
                admin = bool(getattr(perms, "administrator", False))
                manage_guild = bool(getattr(perms, "manage_guild", False))
                manage_messages = bool(getattr(perms, "manage_messages", False))
        except Exception:
            pass

    allowed = False
    if admin or manage_guild or manage_messages:
        allowed = True

    if not allowed and STAFF_ROLE_IDS and role_ids:
        allowed = any((rid in STAFF_ROLE_IDS) for rid in role_ids)

    debug = {
        "user_id": int(interaction.user.id) if interaction.user else None,
        "guild_id": int(interaction.guild_id) if interaction.guild_id else None,
        "detected_role_ids": sorted(role_ids),
        "configured_staff_role_ids": sorted(list(STAFF_ROLE_IDS)),
        "admin": admin,
        "manage_guild": manage_guild,
        "manage_messages": manage_messages,
    }
    return allowed, debug


def staff_only():
    async def predicate(interaction: discord.Interaction) -> bool:
        ok, dbg = await is_staff(interaction)
        if not ok:
            msg = (
                "You do not have permission to run this staff command.\n\n"
                "This bot is configured for **role-based staff access only**.\n"
                "If you expect access, verify **STAFF_ROLE_IDS** contains your staff role IDs, "
                "and ensure the bot has **Server Members Intent** enabled.\n\n"
                f"--- Debug ---\n"
                f"Your user_id: {dbg['user_id']}\n"
                f"Detected role IDs: {dbg['detected_role_ids']}\n"
                f"Configured STAFF_ROLE_IDS (effective): {dbg.get('configured_staff_role_ids')}\n"
                f"Guild ID (interaction): {dbg['guild_id']}\n"
            )
            # best-effort ephemeral response
            try:
                if interaction.response.is_done():
                    await interaction.followup.send(msg, ephemeral=True)
                else:
                    await interaction.response.send_message(msg, ephemeral=True)
            except Exception:
                pass
        return ok
    return app_commands.check(predicate)


# -------------------------
# Characters (from Postgres "characters")
# -------------------------

async def character_autocomplete(
    interaction: discord.Interaction,
    current: str,
) -> List[app_commands.Choice[str]]:
    # Pull from Postgres characters table using DATA_GUILD_ID partition
    like = f"%{(current or '').lower()}%"
    rows = await db_fetch(
        """
        SELECT name
        FROM characters
        WHERE guild_id=$1 AND archived=FALSE AND LOWER(name) LIKE $2
        ORDER BY name ASC
        LIMIT 25;
        """,
        DATA_GUILD_ID,
        like,
    )
    return [app_commands.Choice(name=r["name"], value=r["name"]) for r in rows]



async def ac_asset_for_character(interaction: discord.Interaction, current: str) -> List[app_commands.Choice[str]]:
    """Autocomplete assets for the selected character. Format: 'Type | Tier | Name'."""
    try:
        char = getattr(interaction.namespace, "character", None)
        if not char:
            return []
        rows = await get_assets_for_character(str(char))
        needle = (current or "").lower()
        out: List[app_commands.Choice[str]] = []
        for r in rows:
            if is_noble_title_asset_type(str(r.get("asset_type", ""))):
                continue
            label = f"{r['asset_type']} | {r['tier']} | {r['asset_name']}"
            if needle and needle not in label.lower():
                continue
            out.append(app_commands.Choice(name=label[:100], value=label[:100]))
            if len(out) >= 25:
                break
        return out
    except Exception:
        return []


async def ac_target_tier(interaction: discord.Interaction, current: str) -> List[app_commands.Choice[str]]:
    """Autocomplete higher tiers for the selected asset's type."""
    try:
        asset_label = getattr(interaction.namespace, "asset", None)
        if not asset_label:
            return []
        parts = [p.strip() for p in str(asset_label).split("|")]
        if len(parts) < 3:
            return []
        asset_type = parts[0]
        cur_tier = parts[1]
        cur_rank = _tier_rank(cur_tier)

        tiers = await list_tiers_for_type(asset_type)
        needle = (current or "").lower()
        out: List[app_commands.Choice[str]] = []
        for t in tiers:
            tr = _tier_rank(t)
            if cur_rank is not None and tr is not None and tr <= cur_rank:
                continue
            if needle and needle not in t.lower():
                continue
            out.append(app_commands.Choice(name=t[:100], value=t[:100]))
            if len(out) >= 25:
                break
        return out
    except Exception:
        return []


async def noble_title_family_autocomplete(interaction: discord.Interaction, current: str) -> List[app_commands.Choice[str]]:
    needle = (current or "").lower()
    out: List[app_commands.Choice[str]] = []
    for family in NOBLE_TITLE_FAMILIES.keys():
        if needle and needle not in family.lower():
            continue
        out.append(app_commands.Choice(name=family, value=family))
        if len(out) >= 25:
            break
    return out


async def noble_title_option_autocomplete(interaction: discord.Interaction, current: str) -> List[app_commands.Choice[str]]:
    family = str(getattr(interaction.namespace, "title_name", "") or getattr(interaction.namespace, "new_title_name", "") or "").strip()
    if not family or family not in NOBLE_TITLE_FAMILIES:
        return []
    needle = (current or "").lower()
    out: List[app_commands.Choice[str]] = []
    for opt in NOBLE_TITLE_FAMILIES[family]["options"]:
        if needle and needle not in opt.lower():
            continue
        out.append(app_commands.Choice(name=opt, value=opt))
        if len(out) >= 25:
            break
    return out


async def noble_title_realm_autocomplete(interaction: discord.Interaction, current: str) -> List[app_commands.Choice[str]]:
    try:
        char = getattr(interaction.namespace, "character", None)
        if not char:
            return []
        rows = await get_noble_titles_for_character(str(char))
        needle = (current or "").lower()
        out: List[app_commands.Choice[str]] = []
        for r in rows:
            label = noble_title_display_from_row(r)
            ident = noble_title_identifier_from_row(r)
            if needle and needle not in label.lower() and needle not in ident.lower():
                continue
            out.append(app_commands.Choice(name=label[:100], value=ident[:100]))
            if len(out) >= 25:
                break
        return out
    except Exception:
        return []


async def get_character_owner(character_name: str) -> Optional[int]:
    row = await db_fetchrow(
        """
        SELECT user_id
        FROM characters
        WHERE guild_id=$1 AND name=$2 AND archived=FALSE
        LIMIT 1;
        """,
        DATA_GUILD_ID,
        character_name,
    )
    return int(row["user_id"]) if row else None



async def get_character_kingdom(character_name: str) -> Optional[str]:
    row = await db_fetchrow(
        """
        SELECT kingdom
        FROM characters
        WHERE guild_id=$1 AND name=$2 AND archived=FALSE
        LIMIT 1;
        """,
        DATA_GUILD_ID,
        character_name,
    )
    if not row:
        return None
    k = row.get("kingdom")
    if k is None:
        return None
    k = str(k).strip()
    return k if k else None


# -------------------------
# Economy core helpers
# -------------------------

async def get_balance(character_name: str) -> int:
    row = await db_fetchrow(
        """
        SELECT balance_val
        FROM econ_balances
        WHERE guild_id=$1 AND character_name=$2
        LIMIT 1;
        """,
        DATA_GUILD_ID,
        character_name,
    )
    return int(row["balance_val"]) if row else 0


async def set_balance(character_name: str, new_val: int) -> None:
    await db_exec(
        """
        INSERT INTO econ_balances (guild_id, character_name, balance_val, updated_at)
        VALUES ($1, $2, $3, NOW())
        ON CONFLICT (guild_id, character_name)
        DO UPDATE SET balance_val=EXCLUDED.balance_val, updated_at=NOW();
        """,
        DATA_GUILD_ID,
        character_name,
        int(new_val),
    )


async def adjust_balance(character_name: str, delta: int) -> int:
    cur = await get_balance(character_name)
    new_val = cur + int(delta)
    await set_balance(character_name, new_val)
    return new_val


async def log_audit(interaction: discord.Interaction, action: str, details: Dict[str, Any]) -> None:
    try:
        await db_exec(
            """
            INSERT INTO econ_audit_log (ts, guild_id, actor_user_id, action, details)
            VALUES (NOW(), $1, $2, $3, $4::jsonb);
            """,
            DATA_GUILD_ID,
            int(interaction.user.id),
            action,
            json.dumps(details),
        )
    except Exception as e:
        print(f"[warn] audit log insert failed: {e}")


async def _safe_actor_name(interaction: discord.Interaction) -> str:
    # Never mention/ping users. Prefer server nickname (display_name) when available.
    try:
        u = interaction.user
        return getattr(u, "display_name", None) or getattr(u, "name", "Unknown")
    except Exception:
        return "Unknown"


def _fmt_kv(details: Dict[str, Any], keys: List[str]) -> List[str]:
    lines: List[str] = []
    for k in keys:
        if k in details and details[k] is not None and details[k] != "":
            lines.append(f"- **{k.replace('_', ' ').title()}:** {details[k]}")
    return lines


async def log_econ_channel(interaction: discord.Interaction, action: str, details: Dict[str, Any]) -> None:
    """Send an econ action log line to the configured Discord channel (no mentions)."""
    if not ECON_LOG_CHANNEL_ID:
        return
    try:
        channel = client.get_channel(int(ECON_LOG_CHANNEL_ID))
        if channel is None:
            channel = await client.fetch_channel(int(ECON_LOG_CHANNEL_ID))  # type: ignore
        if channel is None:
            print(f"[warn] ECON log channel not found: {ECON_LOG_CHANNEL_ID}")
            return

        actor_name = await _safe_actor_name(interaction)
        actor_id = int(interaction.user.id) if getattr(interaction, "user", None) else 0

        header = f"**ECON LOG:** `{action}`"
        actor_line = f"**Actor:** {actor_name} ({actor_id})"

        lines: List[str] = [header, actor_line]

        if action == "purchase_new":
            lines += _fmt_kv(details, ["character", "tier", "asset_name", "asset_type", "cost", "add_income", "sales_kingdom", "new_balance"])
        elif action == "upgrade_asset":
            lines += _fmt_kv(details, ["character", "asset_name", "asset_type", "from_tier", "to_tier", "cost", "sales_kingdom"])
        elif action == "sell_asset":
            lines += _fmt_kv(details, ["character", "asset_name", "asset_type", "tier", "gross_refund", "tax_amount", "net_refund"])
        elif action == "income_claim":
            lines += _fmt_kv(details, ["character", "character_kingdom", "base_income", "asset_income", "gross_total", "tax_total", "net_total", "new_balance"])
        elif action in ("adjust_balance", "set_balance"):
            lines += _fmt_kv(details, ["character", "delta", "value", "new_balance"])
        elif action in ("set_kingdom_tax", "set_kingdom_treasury"):
            lines += _fmt_kv(details, ["kingdom", "percent", "tax_rate_bp", "treasury"])
        elif action == "assign_noble_title":
            lines += _fmt_kv(details, ["character", "title", "kingdom"])
        else:
            lines += _fmt_kv(details, ["character", "kingdom", "amount", "cost", "new_balance"])

        msg = _cap_message("\n".join(lines))
        await channel.send(msg, allowed_mentions=discord.AllowedMentions.none())  # type: ignore
    except Exception as e:
        print(f"[warn] ECON channel log failed: {e}")


async def log_econ(interaction: discord.Interaction, action: str, details: Dict[str, Any]) -> None:
    """Write to DB audit log and to Discord econ log channel (if configured).

    IMPORTANT: This must never call itself (no recursion). It wraps:
      - log_audit (DB json audit)
      - log_econ_channel (Discord channel log)
    """
    await log_audit(interaction, action, details)
    await log_econ_channel(interaction, action, details)



# -------------------------
# Assets (definitions in econ_asset_definitions; purchases in econ_assets)
# -------------------------

async def list_asset_types() -> List[str]:
    try:
        rows = await db_fetch(
            """
            SELECT DISTINCT asset_type
            FROM econ_asset_definitions
            WHERE asset_type <> 'Noble Title'
            ORDER BY asset_type ASC;
            """
        )
        return [str(r["asset_type"]) for r in rows]
    except Exception as e:
        print(f"[warn] list_asset_types failed: {e}")
        return []


async def list_tiers_for_type(asset_type: str) -> List[str]:
    try:
        rows = await db_fetch(
            """
            SELECT tier
            FROM econ_asset_definitions
            WHERE asset_type=$1
            ORDER BY tier ASC;
            """,
            asset_type,
        )
        return [str(r["tier"]) for r in rows]
    except Exception as e:
        print(f"[warn] list_tiers_for_type({asset_type}) failed: {e}")
        return []


async def get_asset_def(asset_type: str, tier: str) -> Optional[Tuple[int, int]]:
    row = await db_fetchrow(
        """
        SELECT cost_val, add_income_val
        FROM econ_asset_definitions
        WHERE asset_type=$1 AND tier=$2
        LIMIT 1;
        """,
        asset_type,
        tier,
    )
    if not row:
        return None
    return int(row["cost_val"]), int(row["add_income_val"])





async def get_asset_definition_kingdom(asset_type: str, tier: str) -> Optional[str]:
    """Return the sales/tax kingdom assigned to this asset definition row (nullable)."""
    try:
        row = await db_fetchrow(
            """
            SELECT kingdom
            FROM econ_asset_definitions
            WHERE asset_type=$1 AND tier=$2
            LIMIT 1;
            """,
            asset_type,
            tier,
        )
        if not row:
            return None
        k = row.get("kingdom")
        if k is None:
            return None
        k = str(k).strip()
        return k if k else None
    except Exception:
        return None

def _selected_option_from_interaction(interaction: discord.Interaction, option_name: str) -> Optional[str]:
    """
    Robustly retrieve a selected option value during autocomplete.
    discord.py sometimes lacks namespace fields during autocomplete depending on client/event.
    """
    # 1) Try namespace
    try:
        ns = getattr(interaction, "namespace", None)
        if ns is not None and hasattr(ns, option_name):
            v = getattr(ns, option_name)
            if v is not None:
                return str(v)
    except Exception:
        pass
    # 2) Try interaction.data options payload
    try:
        data = getattr(interaction, "data", None) or {}
        opts = data.get("options") or []
        # options can be nested for groups; handle shallow only (we don't use groups here)
        for o in opts:
            if o.get("name") == option_name and "value" in o:
                return str(o.get("value"))
    except Exception:
        pass
    return None
async def seed_asset_definitions() -> None:
    """Upsert the authoritative asset definitions set into econ_asset_definitions."""
    try:
        await db_exec(
            """
            CREATE TABLE IF NOT EXISTS econ_asset_definitions (
              asset_type TEXT NOT NULL,
              tier TEXT NOT NULL,
              cost_val BIGINT NOT NULL,
              add_income_val BIGINT NOT NULL,
              kingdom TEXT,
              PRIMARY KEY (asset_type, tier)
            );
            """
        )
        # Additive migration: asset sales kingdom (nullable)
        await db_exec("ALTER TABLE econ_asset_definitions ADD COLUMN IF NOT EXISTS kingdom TEXT;")
        for asset_type, tier, cost_val, add_income_val in ASSET_DEFINITIONS_SEED:
            await db_exec(
                """
                INSERT INTO econ_asset_definitions (asset_type, tier, cost_val, add_income_val)
                VALUES ($1, $2, $3, $4)
                ON CONFLICT (asset_type, tier)
                DO NOTHING;
                """,
                asset_type, tier, int(cost_val), int(add_income_val)
            )
        row = await db_fetchrow("SELECT COUNT(*) AS c FROM econ_asset_definitions;")
        total = int(row["c"]) if row and "c" in row else 0
        if total != len(ASSET_DEFINITIONS_SEED):
            print(f"[warn] econ_asset_definitions rowcount={total} differs from seed={len(ASSET_DEFINITIONS_SEED)}.")
        else:
            print(f"[test] econ_asset_definitions seeded/verified: {total} row(s).")
    except Exception as e:
        print(f"[warn] seed_asset_definitions failed: {e}")



def _parse_val_cell(v: Any) -> Optional[int]:
    """Parse a spreadsheet cell that may look like '300 Val' or 300 into an int (base Val/Cinth units)."""
    if v is None:
        return None
    if isinstance(v, (int,)):
        return int(v)
    if isinstance(v, float):
        # Spreadsheet might store whole numbers as floats
        return int(round(v))
    s = str(v).strip()
    if not s:
        return None
    # Extract the first integer-like token
    m = re.search(r"-?\d+", s.replace(",", ""))
    if not m:
        return None
    try:
        return int(m.group(0))
    except Exception:
        return None


async def import_asset_definitions_from_xlsx_bytes(data: bytes) -> Tuple[int, int, int, List[str]]:
    """Import/update econ_asset_definitions from an uploaded NEW Asset Table.xlsx.

    Policy: UPSERT by (asset_type, tier). Existing rows are UPDATED (no duplicates).
    Returns: (rows_processed, inserted_count, updated_count, errors)
    """
    errors: List[str] = []
    if openpyxl is None:
        return 0, 0, 0, ["openpyxl is not available in this runtime."]

    try:
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    except Exception as e:
        return 0, 0, 0, [f"Failed to read XLSX: {e}"]

    # Use first worksheet
    ws = wb.worksheets[0]

    # Read header row
    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {str(h).strip(): i for i, h in enumerate(header_row) if h is not None}

    required = ["Asset Type", "Tier", "Cost to Acquire", "Add to Income"]
    # Optional (but recommended): Kingdom (destination for purchase/upgrade funds + income tax bucket)
    i_kingdom = header_map.get("Kingdom")
    missing = [h for h in required if h not in header_map]
    if missing:
        return 0, 0, 0, [f"XLSX missing required column(s): {', '.join(missing)}"]

    i_type = header_map["Asset Type"]
    i_tier = header_map["Tier"]
    i_cost = header_map["Cost to Acquire"]
    i_inc = header_map["Add to Income"]

    processed = 0
    inserted = 0
    updated = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue

        asset_type = str(row[i_type]).strip() if row[i_type] is not None else ""
        tier = str(row[i_tier]).strip() if row[i_tier] is not None else ""
        cost_val = _parse_val_cell(row[i_cost])
        add_income_val = _parse_val_cell(row[i_inc]) or 0
        kingdom_val: Optional[str] = None
        if i_kingdom is not None and i_kingdom < len(row):
            kv = row[i_kingdom]
            if kv is not None:
                k = str(kv).strip()
                if k:
                    # Validate against canonical list to avoid typos silently creating new kingdoms.
                    if k not in CANON_KINGDOMS:
                        errors.append(f"Row {row_idx}: invalid Kingdom '{k}' (must be one of: {', '.join(CANON_KINGDOMS)})")
                        continue
                    kingdom_val = k

        if not asset_type or not tier:
            errors.append(f"Row {row_idx}: missing Asset Type or Tier")
            continue
        if cost_val is None:
            errors.append(f"Row {row_idx}: could not parse Cost to Acquire")
            continue

        processed += 1

        # UPSERT; determine insert vs update using xmax=0 trick
        rec = await db_fetchrow(
            """
            WITH upsert AS (
              INSERT INTO econ_asset_definitions(asset_type, tier, cost_val, add_income_val, kingdom)
              VALUES ($1, $2, $3, $4, $5)
              ON CONFLICT (asset_type, tier)
              DO UPDATE SET cost_val=EXCLUDED.cost_val, add_income_val=EXCLUDED.add_income_val,
                            kingdom=COALESCE(EXCLUDED.kingdom, econ_asset_definitions.kingdom)
              RETURNING (xmax = 0) AS inserted
            )
            SELECT inserted FROM upsert;
            """,
            asset_type,
            tier,
            int(cost_val),
            int(add_income_val),
            kingdom_val,
        )
        if rec and bool(rec["inserted"]):
            inserted += 1
        else:
            updated += 1

    return processed, inserted, updated, errors


async def asset_type_autocomplete(interaction: discord.Interaction, current: str) -> List[app_commands.Choice[str]]:
    current_l = (current or "").lower()
    types = await list_asset_types()
    if not types:
        print('[warn] econ_asset_definitions returned 0 asset types (table empty or not populated).')
    out = [t for t in types if not is_noble_title_asset_type(t) and current_l in t.lower()][:25]
    return [app_commands.Choice(name=t, value=t) for t in out]


async def tier_autocomplete(interaction: discord.Interaction, current: str) -> List[app_commands.Choice[str]]:
    # Need selected asset_type to filter tiers
    asset_type = _selected_option_from_interaction(interaction, "asset_type") or ""
    if not asset_type:
        return []
    tiers = await list_tiers_for_type(asset_type)
    current_l = (current or "").lower()
    out = [t for t in tiers if current_l in t.lower()][:25]
    return [app_commands.Choice(name=t, value=t) for t in out]


async def recompute_daily_income(character_name: str) -> int:
    # Daily income = sum(add_income_val) across current assets' tiers
    rows = await db_fetch(
        """
        SELECT a.asset_type, a.tier, d.add_income_val
        FROM econ_assets a
        JOIN econ_asset_definitions d
          ON d.asset_type=a.asset_type AND d.tier=a.tier
        WHERE a.guild_id=$1 AND a.character_name=$2;
        """,
        DATA_GUILD_ID,
        character_name,
    )
    return sum(int(r["add_income_val"]) for r in rows)


# -------------------------
# Kingdom taxation helpers
# -------------------------

def _bp_to_percent(bp: int) -> str:
    # 100 bp = 1%
    return f"{bp / 100:.0f}%" if bp % 100 == 0 else f"{bp / 100:.2f}%"

def _calc_tax(amount_cinth: int, tax_rate_bp: int) -> int:
    # Whole-cinth rule: ALWAYS round DOWN (floor) to nearest cinth.
    if amount_cinth <= 0 or tax_rate_bp <= 0:
        return 0
    return (int(amount_cinth) * int(tax_rate_bp)) // 10000

async def get_character_kingdom(character_name: str) -> Optional[str]:
    # Home kingdom lives in characters table (populated by upstream bot).
    try:
        row = await db_fetchrow(
            """
            SELECT kingdom
            FROM characters
            WHERE guild_id=$1 AND name=$2 AND archived=FALSE
            LIMIT 1;
            """,
            DATA_GUILD_ID,
            character_name,
        )
        if not row:
            return None
        hk = row.get("kingdom")
        return str(hk).strip() if hk is not None and str(hk).strip() else None
    except Exception:
        return None

async def get_kingdom_tax_bp(kingdom: str) -> int:
    row = await db_fetchrow(
        """
        SELECT tax_rate_bp
        FROM econ_kingdoms
        WHERE guild_id=$1 AND kingdom=$2
        LIMIT 1;
        """,
        DATA_GUILD_ID,
        kingdom,
    )
    return int(row["tax_rate_bp"]) if row else 0

async def upsert_kingdom_tax_bp(kingdom: str, tax_rate_bp: int) -> None:
    await db_exec(
        """
        INSERT INTO econ_kingdoms (guild_id, kingdom, tax_rate_bp, treasury)
        VALUES ($1, $2, $3, 0)
        ON CONFLICT (guild_id, kingdom)
        DO UPDATE SET tax_rate_bp=EXCLUDED.tax_rate_bp;
        """,
        DATA_GUILD_ID,
        kingdom,
        int(tax_rate_bp),
    )

async def add_to_kingdom_treasury(kingdom: str, amount_cinth: int) -> None:
    if amount_cinth <= 0:
        return
    await db_exec(
        """
        INSERT INTO econ_kingdoms (guild_id, kingdom, tax_rate_bp, treasury)
        VALUES ($1, $2, 0, $3)
        ON CONFLICT (guild_id, kingdom)
        DO UPDATE SET treasury=econ_kingdoms.treasury + EXCLUDED.treasury;
        """,
        DATA_GUILD_ID,
        kingdom,
        int(amount_cinth),
    )



async def set_kingdom_treasury_exact(kingdom: str, treasury_cinth: int) -> None:
    """Set a kingdom treasury to an exact value (cinths). Preserves existing tax_rate_bp."""
    treasury_cinth = int(treasury_cinth)
    if treasury_cinth < 0:
        treasury_cinth = 0
    await db_exec(
        """
        INSERT INTO econ_kingdoms (guild_id, kingdom, tax_rate_bp, treasury)
        VALUES ($1, $2, 0, $3)
        ON CONFLICT (guild_id, kingdom)
        DO UPDATE SET treasury=EXCLUDED.treasury;
        """,
        DATA_GUILD_ID,
        kingdom,
        treasury_cinth,
    )
async def fetch_kingdom_treasuries() -> List[Tuple[str, int, int]]:
    rows = await db_fetch(
        """
        SELECT kingdom, tax_rate_bp, treasury
        FROM econ_kingdoms
        WHERE guild_id=$1
        ORDER BY kingdom ASC;
        """,
        DATA_GUILD_ID,
    )
    out: List[Tuple[str, int, int]] = []
    for r in rows:
        out.append((str(r["kingdom"]), int(r["tax_rate_bp"]), int(r["treasury"])))
    return out

async def render_treasury_lines() -> List[str]:
    treas = await fetch_kingdom_treasuries()
    if not treas:
        return [
            "🏰 **Kingdom Treasuries**",
            "━━━━━━━━━━━━━━━━━━",
            "_No kingdoms configured yet._",
            "",
        ]
    out: List[str] = [
        "🏰 **Kingdom Treasuries**",
        "━━━━━━━━━━━━━━━━━━",
    ]
    for kingdom, bp, treasury in treas:
        out.append(f"• **{kingdom}** — Treasury: **{format_currency(treasury)}** — Tax: **{_bp_to_percent(bp)}**")
    out.append("")
    return out



# -------------------------
# Bank dashboard persistence (approved)
# -------------------------

async def bank_message_ids_from_db() -> List[int]:
    rows = await db_fetch(
        """
        SELECT idx, message_id
        FROM econ_bank_messages
        WHERE guild_id=$1
        ORDER BY idx ASC;
        """,
        DATA_GUILD_ID,
    )
    if not rows:
        return []
    # fill by idx order
    return [int(r["message_id"]) for r in rows]


async def save_bank_message_ids(message_ids: List[int]) -> None:
    # Upsert by idx
    pool = await db_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            # delete existing for guild, then insert (keeps idx stable)
            await conn.execute("DELETE FROM econ_bank_messages WHERE guild_id=$1;", DATA_GUILD_ID)
            for i, mid in enumerate(message_ids):
                await conn.execute(
                    """
                    INSERT INTO econ_bank_messages (guild_id, idx, message_id)
                    VALUES ($1, $2, $3);
                    """,
                    DATA_GUILD_ID,
                    int(i),
                    int(mid),
                )


async def fetch_bank_dashboard_rows() -> List[Tuple[str, int, int, int]]:
    """rows: [(character_name, owner_id, balance, income)]"""
    chars = await db_fetch(
        """
        SELECT user_id, name
        FROM characters
        WHERE guild_id=$1 AND archived=FALSE
        ORDER BY name ASC;
        """,
        DATA_GUILD_ID,
    )

    rows: List[Tuple[str, int, int, int]] = []
    for r in chars:
        cname = str(r["name"])
        uid = int(r["user_id"])
        bal = await get_balance(cname)
        inc = await recompute_daily_income(cname)
        rows.append((cname, uid, bal, inc))
    return rows


def _truncate_embed_field(text: str, limit: int = 1024) -> str:
    text = str(text or "").strip()
    if len(text) <= limit:
        return text or "-"
    if limit <= 3:
        return text[:limit]
    return text[: limit - 3].rstrip() + "..."


def _bank_color_for_kingdom(kingdom: str) -> int:
    k = str(kingdom or "").strip().lower()
    palette = {
        "velarith": 0x7F98B1,
        "lyvik": 0x6B8E9E,
        "avalea": 0xB7C9E2,
        "baelon": 0x8C5A4B,
        "sethrathiel": 0x6E5A8A,
        "setrathiel": 0x6E5A8A,
    }
    return palette.get(k, 0x8B7355)


async def _sorted_bank_dashboard_rows_for_display(guild: discord.Guild) -> List[Tuple[str, int, int, int, str]]:
    """rows: [(character_name, owner_id, balance, income, owner_display)] grouped by owner display then character."""
    rows = await fetch_bank_dashboard_rows()
    decorated: List[Tuple[str, int, int, int, str]] = []
    for cname, uid, bal, inc in rows:
        owner_display = sanitize_plain_text(await _display_name_from_cache(guild, int(uid or 0))) if uid else "Unknown Owner"
        decorated.append((cname, uid, bal, inc, owner_display))
    decorated.sort(key=lambda r: (str(r[4]).lower(), str(r[0]).lower()))
    return decorated


async def render_bank_header_embed(guild: discord.Guild) -> discord.Embed:
    rows = await _sorted_bank_dashboard_rows_for_display(guild)
    now = datetime.now(CHICAGO_TZ)

    total_balance = sum(int(r[2]) for r in rows)
    total_income = sum(int(r[3]) for r in rows)
    owner_count = len({int(r[1]) for r in rows if int(r[1] or 0) > 0})

    embed = discord.Embed(
        title="🏦 Bank of Vilyra",
        description="The crown ledger of coin, titles, and holdings across the realm.",
        color=0x7F98B1,
    )

    treas = await fetch_kingdom_treasuries()
    treasury_lines: List[str] = []
    for kingdom, bp, treasury in treas:
        treasury_lines.append(
            f"👑 **{sanitize_plain_text(kingdom)}** — {format_currency(int(treasury or 0))} • Tax {int(bp or 0) / 100:.0f}%"
        )
    embed.add_field(
        name="🪙 Kingdom Treasuries",
        value=_truncate_embed_field("\n".join(treasury_lines) or "(none)"),
        inline=False,
    )

    if rows:
        top_bal = sorted(rows, key=lambda r: int(r[2]), reverse=True)[:5]
        top_inc = sorted(rows, key=lambda r: int(r[3]), reverse=True)[:5]

        bal_lines = [
            f"{i}. **{sanitize_plain_text(c)}** — {format_currency(bal)}"
            for i, (c, _uid, bal, _inc, _owner) in enumerate(top_bal, start=1)
        ]
        inc_lines = [
            f"{i}. **{sanitize_plain_text(c)}** — {format_currency(inc)}"
            for i, (c, _uid, _bal, inc, _owner) in enumerate(top_inc, start=1)
        ]

        embed.add_field(name="💰 Wealthiest Ledgers", value=_truncate_embed_field("\n".join(bal_lines) or "(none)"), inline=False)
        embed.add_field(name="📈 Highest Daily Income", value=_truncate_embed_field("\n".join(inc_lines) or "(none)"), inline=False)
        embed.add_field(
            name="📜 Ledger Status",
            value=_truncate_embed_field(
                "\n".join([
                    f"Characters tracked: **{len(rows)}**",
                    f"Owners represented: **{owner_count}**",
                    f"Total wealth: **{format_currency(total_balance)}**",
                    f"Total daily income: **{format_currency(total_income)}**",
                    "Character ledgers appear below, grouped by owner.",
                ])
            ),
            inline=False,
        )
    else:
        embed.add_field(name="📜 Ledger Status", value="No characters found in DB.", inline=False)

    embed.set_footer(text=f"Last updated {now.strftime('%Y-%m-%d %H:%M')} America/Chicago")
    return embed


async def render_owner_group_embed(guild: discord.Guild, owner_id: int, owner_display: str, character_names: List[str]) -> discord.Embed:
    clean_owner = sanitize_plain_text(owner_display or f"User {owner_id}")
    lines = [f"• {sanitize_plain_text(name)}" for name in character_names]
    embed = discord.Embed(
        title=f"👤 {clean_owner}",
        description="Household ledger grouping",
        color=0x5C6F82,
    )
    embed.add_field(
        name="Characters",
        value=_truncate_embed_field("\n".join(lines) or "(none)"),
        inline=False,
    )
    embed.set_footer(text="The following ledgers belong to this owner")
    return embed


async def render_character_bank_embed(guild: discord.Guild, character_name: str, embed_color: Optional[int] = None) -> discord.Embed:
    owner_id = await get_character_owner(character_name)
    owner_display = sanitize_plain_text(await _display_name_from_cache(guild, int(owner_id or 0))) if owner_id else "Unknown Owner"
    kingdom = sanitize_plain_text(str(await get_character_kingdom(character_name) or "")).strip() or "(No Kingdom)"
    balance_val = await get_balance(character_name)
    assets = await get_assets_with_income_for_character(character_name)
    asset_income_sum = sum(int(a.get("add_income_val") or 0) for a in assets)
    base_income = int(current_base_daily_income())
    total_income = base_income + asset_income_sum
    title_assets = [a for a in assets if is_noble_title_asset_type(str(a.get("asset_type", "")))]
    regular_assets = [a for a in assets if not is_noble_title_asset_type(str(a.get("asset_type", "")))]

    embed = discord.Embed(
        title=f"📘 {sanitize_plain_text(character_name)}",
        description=f"Held under **{owner_display}**",
        color=(int(embed_color) if embed_color is not None else _bank_color_for_kingdom(kingdom)),
    )
    embed.add_field(name="🏰 Kingdom", value=kingdom, inline=True)
    embed.add_field(name="💰 Balance", value=_truncate_embed_field(format_balance(balance_val), 1024), inline=False)
    embed.add_field(name="☀️ Base Daily Income", value=format_currency(base_income), inline=True)
    embed.add_field(name="🏦 Asset Income", value=format_currency(asset_income_sum), inline=True)
    embed.add_field(name="📈 Total Daily Income", value=format_currency(total_income), inline=True)

    if title_assets:
        title_lines = [f"👑 {noble_title_display_from_row(a)}" for a in title_assets]
        embed.add_field(name="👑 Noble Titles", value=_truncate_embed_field("\n".join(title_lines), 1024), inline=False)
    else:
        embed.add_field(name="👑 Noble Titles", value="None recorded", inline=False)

    if regular_assets:
        asset_lines: List[str] = []
        hidden = 0
        for a in regular_assets:
            tier = _tier_label(str(a.get("tier", "")))
            aname = sanitize_plain_text(str(a.get("asset_name", "")).strip() or str(a.get("asset_type", "")).strip())
            akingdom = sanitize_plain_text(str(a.get("kingdom", "")).strip() or "(No Kingdom)")
            add = int(a.get("add_income_val") or 0)
            candidate = f"• **{tier}** — {aname} • {akingdom} • +{add:,} Val"
            joined = "\n".join(asset_lines + [candidate])
            if len(joined) > 1000:
                hidden += 1
            else:
                asset_lines.append(candidate)
        if hidden > 0:
            asset_lines.append(f"• ...and {hidden} more")
        embed.add_field(name="🧾 Assets", value=_truncate_embed_field("\n".join(asset_lines), 1024), inline=False)
    else:
        embed.add_field(name="🧾 Assets", value="No income-bearing assets recorded", inline=False)

    embed.set_footer(text=f"Live ledger • Updated {datetime.now(CHICAGO_TZ).strftime('%Y-%m-%d %H:%M')} America/Chicago")
    return embed


async def render_bank_dashboard_embeds(guild: discord.Guild) -> List[discord.Embed]:
    rows = await _sorted_bank_dashboard_rows_for_display(guild)
    embeds: List[discord.Embed] = [await render_bank_header_embed(guild)]
    grouped: Dict[int, Dict[str, Any]] = {}
    order: List[int] = []
    for cname, uid, _bal, _inc, owner_display in rows:
        key = int(uid or 0)
        if key not in grouped:
            grouped[key] = {"owner_display": owner_display, "characters": []}
            order.append(key)
        grouped[key]["characters"].append(cname)

    owner_group_colors = [0x7F98B1, 0x5E748C]
    owner_color_map: Dict[int, int] = {
        uid: owner_group_colors[idx % len(owner_group_colors)]
        for idx, uid in enumerate(order)
    }

    for uid in order:
        char_names = list(grouped[uid]["characters"])
        group_color = owner_color_map.get(uid, owner_group_colors[0])
        for cname in char_names:
            embeds.append(await render_character_bank_embed(guild, cname, embed_color=group_color))
    return embeds


async def refresh_bank_dashboard(create_missing: bool = True, header_only: bool = False) -> None:
    if not BANK_CHANNEL_ID:
        return
    ch = client.get_channel(int(BANK_CHANNEL_ID))
    if ch is None or not isinstance(ch, (discord.TextChannel, discord.Thread)):
        return

    mids = await bank_message_ids_from_db()
    if not mids and BANK_MESSAGE_IDS:
        mids = list(BANK_MESSAGE_IDS)

    embeds = [await render_bank_header_embed(ch.guild)] if header_only else await render_bank_dashboard_embeds(ch.guild)
    if not embeds:
        embeds = [discord.Embed(title="Bank of Vilyra", description="(empty)")]

    msgs: List[discord.Message] = []
    for mid in mids:
        try:
            m = await ch.fetch_message(int(mid))
            msgs.append(m)
        except Exception:
            pass

    if create_missing and len(msgs) < len(embeds):
        try:
            while len(msgs) < len(embeds):
                m = await ch.send(embed=embeds[len(msgs)], allowed_mentions=discord.AllowedMentions.none())
                msgs.append(m)
                await asyncio.sleep(0.35)
            await save_bank_message_ids([int(m.id) for m in msgs])
            print(f"[test] Bank dashboard message IDs saved to Postgres: {len(msgs)}")
        except Exception as e:
            print(f"[warn] Bank dashboard create/persist failed: {e}")

    if header_only:
        if not msgs:
            return
        try:
            await msgs[0].edit(content=None, embed=embeds[0], allowed_mentions=discord.AllowedMentions.none())
        except discord.HTTPException as e:
            print(f"[warn] Failed to edit header page: {e}")
        return

    n = min(len(msgs), len(embeds))
    for i in range(n):
        for attempt in range(4):
            try:
                await msgs[i].edit(content=None, embed=embeds[i], allowed_mentions=discord.AllowedMentions.none())
                if i < n - 1:
                    await asyncio.sleep(0.75)
                break
            except discord.HTTPException:
                await asyncio.sleep(0.8 * (attempt + 1))
            except Exception:
                await asyncio.sleep(0.5 * (attempt + 1))

    if len(msgs) > len(embeds):
        for j in range(len(embeds), len(msgs)):
            try:
                await msgs[j].edit(content="(unused bank slot)", embed=None, allowed_mentions=discord.AllowedMentions.none())
            except Exception:
                pass

# Bank refresh coordinator (Option A1): single worker + dirty flags to avoid overlapping refreshes.
_bank_refresh_task: Optional[asyncio.Task] = None
_bank_refresh_lock: asyncio.Lock = asyncio.Lock()
_bank_dirty_full: bool = False
_bank_dirty_header: bool = False

async def force_rebuild_bank_dashboard() -> None:
    """Delete and recreate all tracked bank dashboard messages from fresh rendered embeds."""
    if not BANK_CHANNEL_ID:
        return
    ch = client.get_channel(int(BANK_CHANNEL_ID))
    if ch is None or not isinstance(ch, (discord.TextChannel, discord.Thread)):
        return

    embeds = await render_bank_dashboard_embeds(ch.guild)
    if not embeds:
        embeds = [discord.Embed(title="Bank of Vilyra", description="(empty)")]

    mids = await bank_message_ids_from_db()
    if not mids and BANK_MESSAGE_IDS:
        mids = list(BANK_MESSAGE_IDS)

    for mid in mids:
        try:
            m = await ch.fetch_message(int(mid))
            await m.delete()
            await asyncio.sleep(0.35)
        except Exception:
            pass

    try:
        await db_exec("DELETE FROM econ_bank_messages WHERE guild_id=$1;", DATA_GUILD_ID)
    except Exception:
        pass

    new_ids = []
    for emb in embeds:
        m = await ch.send(embed=emb, allowed_mentions=discord.AllowedMentions.none())
        new_ids.append(int(m.id))
        await asyncio.sleep(0.35)

    await save_bank_message_ids(new_ids)
    print(f"[test] Force rebuilt bank dashboard embeds: {len(new_ids)}")


def request_bank_refresh(*, full: bool = True) -> None:
    """Schedule a bank dashboard refresh.

    full=True recomputes all pages (header + cards).
    full=False recomputes only page 1 (header/treasuries/leaderboards).
    Multiple calls are coalesced; if a refresh is already running, we mark the desired work as 'dirty'
    so it runs again immediately afterward.
    """
    global _bank_refresh_task, _bank_dirty_full, _bank_dirty_header
    if full:
        _bank_dirty_full = True
    else:
        _bank_dirty_header = True

    if _bank_refresh_task and not _bank_refresh_task.done():
        return

    async def _worker():
        global _bank_dirty_full, _bank_dirty_header
        async with _bank_refresh_lock:
            # Small debounce window to collapse bursts of updates.
            await asyncio.sleep(1.2)
            # Loop until no more dirty work remains.
            while _bank_dirty_full or _bank_dirty_header:
                do_full = _bank_dirty_full
                do_header = _bank_dirty_header and not do_full
                # consume flags
                if do_full:
                    _bank_dirty_full = False
                    _bank_dirty_header = False
                elif do_header:
                    _bank_dirty_header = False

                try:
                    await refresh_bank_dashboard(create_missing=True, header_only=do_header)
                except Exception as e:
                    print(f"[warn] Bank refresh failed: {e}")
                # small spacing between back-to-back cycles; reduces PATCH 429s
                await asyncio.sleep(0.6)

    _bank_refresh_task = asyncio.create_task(_worker())

# Back-compat alias used across commands
def trigger_bank_refresh() -> None:
    request_bank_refresh(full=True)

def trigger_bank_header_refresh() -> None:
    request_bank_refresh(full=False)
# -------------------------
# Commands
# -------------------------

@tree.command(name="balance", description="View a character's current balance.", guild=discord.Object(id=GUILD_ID))
@app_commands.describe(character="Character name")
@app_commands.autocomplete(character=character_autocomplete)
async def cmd_balance(interaction: discord.Interaction, character: str):
    await interaction.response.defer(ephemeral=True)
    guild = interaction.guild
    if guild is None:
        await interaction.followup.send("This command must be used in a server.", ephemeral=True)
        return
    owner_id = await get_character_owner(character)
    if owner_id is None:
        await interaction.followup.send("Character not found in DB.", ephemeral=True, allowed_mentions=discord.AllowedMentions.none())
        return
    card_lines = await render_user_card_block(guild, int(owner_id), [character])
    # render as a single message if possible
    txt = "\n".join(card_lines).strip()
    if len(txt) > 1900:
        # truncate assets if too long for ephemeral message
        trimmed: List[str] = []
        for ln in card_lines:
            if len("\n".join(trimmed + [ln])) > 1800:
                trimmed.append("… (truncated)")
                break
            trimmed.append(ln)
        txt = "\n".join(trimmed).strip()
    await send_ephemeral_with_parchment(interaction, txt)


@tree.command(name="income", description="Claim daily income for a character.", guild=discord.Object(id=GUILD_ID))
@app_commands.describe(character="Character name")
@app_commands.autocomplete(character=character_autocomplete)
async def cmd_income(interaction: discord.Interaction, character: str):
    await interaction.response.defer(ephemeral=True)

    # Ownership: only the character owner can claim (based on characters table user_id)
    owner = await get_character_owner(character)
    if owner is None:
        await interaction.followup.send("Character not found in DB.", ephemeral=True)
        return
    if int(owner) != int(interaction.user.id):
        await interaction.followup.send("You are not the owner of that character.", ephemeral=True)
        return

    today = datetime.now(CHICAGO_TZ).date() if ZoneInfo else date.today()
    row = await db_fetchrow(
        """
        SELECT last_claim_date
        FROM econ_income_claims
        WHERE guild_id=$1 AND character_name=$2
        LIMIT 1;
        """,
        DATA_GUILD_ID,
        character,
    )
    if row and row["last_claim_date"] == today:
        await interaction.followup.send("Daily income already claimed today.", ephemeral=True)
        return

    
    # Kingdom taxation:
    # - Base income is taxed to the character's home kingdom.
    # - Each asset's income is taxed to its own kingdom if set; otherwise inherits home kingdom.
    character_kingdom = await get_character_kingdom(character)
    if not character_kingdom:
        await interaction.followup.send(
            "This character has no **home kingdom** set in the `characters` table. Income cannot be claimed until it is set.",
            ephemeral=True,
        )
        return

    # Pull per-asset incomes so we can bucket taxes per kingdom (whole-cinth only).
    asset_rows = await db_fetch(
        """
        SELECT a.asset_type, a.tier, a.asset_name, COALESCE(a.kingdom, '') AS asset_kingdom, d.add_income_val
        FROM econ_assets a
        JOIN econ_asset_definitions d
          ON d.asset_type=a.asset_type AND d.tier=a.tier
        WHERE a.guild_id=$1 AND a.character_name=$2;
        """,
        DATA_GUILD_ID,
        character,
    )

    asset_income = sum(int(r["add_income_val"]) for r in asset_rows)
    base_income = int(current_base_daily_income())
    gross_total = base_income + int(asset_income or 0)

    # Build kingdom buckets (gross amounts per kingdom)
    buckets: Dict[str, int] = {}
    buckets[character_kingdom] = buckets.get(character_kingdom, 0) + base_income
    for r in asset_rows:
        k = str(r["asset_kingdom"] or "").strip()
        if not k:
            k = character_kingdom
        buckets[k] = buckets.get(k, 0) + int(r["add_income_val"])

    # Compute tax per kingdom bucket (round DOWN) and update treasuries
    total_tax = 0
    for k, amt in buckets.items():
        bp = await get_kingdom_tax_bp(k)
        tax = _calc_tax(int(amt), int(bp))
        if tax > 0:
            await add_to_kingdom_treasury(k, tax)
        total_tax += int(tax)

    net_total = int(gross_total) - int(total_tax)
    if net_total < 0:
        net_total = 0  # safety; should not happen with floor-tax

    # Add NET income to balance
    new_bal = await adjust_balance(character, net_total)

    await db_exec(
        """
        INSERT INTO econ_income_claims (guild_id, character_name, last_claim_date)
        VALUES ($1, $2, $3)
        ON CONFLICT (guild_id, character_name)
        DO UPDATE SET last_claim_date=EXCLUDED.last_claim_date;
        """,
        DATA_GUILD_ID,
        character,
        today,
    )


    await log_econ(
        interaction,
        "income_claim",
        {
            "character": character,
            "character_kingdom": character_kingdom,
            "base_income": base_income,
            "asset_income": asset_income,
            "gross_total": gross_total,
            "tax_total": total_tax,
            "net_total": net_total,
            "new_balance": new_bal,
            "buckets": buckets,
        },
    )

    trigger_bank_refresh()

    msg = (
        f"Claimed daily income for **{character}**:\n"
        f"• Base: **{format_currency(base_income)}** (taxed to **{character_kingdom}**)\n"
        f"• Assets: **{format_currency(asset_income)}**\n"
        f"• Gross: **{format_currency(gross_total)}**\n"
        f"• Tax (rounded down): **{format_currency(total_tax)}**\n"
        f"• Net received: **{format_currency(net_total)}**\n\n"
        f"New balance: **{format_currency(new_bal)}**"
    )
    now_dt = datetime.now(CHICAGO_TZ)
    if now_dt.weekday() in (4, 5, 6):
        await send_ephemeral_with_weekend_income_image(interaction, msg)
    else:
        await send_ephemeral_with_parchment(interaction, msg)




@tree.command(name="econ_commands", description="List EconBot commands.", guild=discord.Object(id=GUILD_ID))
@staff_only()
async def cmd_econ_commands(interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)
    msg = (
        f"**EconBot Commands** ({APP_VERSION})\n\n"
        "**Player**\n"
        "• `/balance` - view a character balance card\n"
        "• `/income` - claim daily income\n\n"
        "**Staff**\n"
        "• `/purchase_new` - record an asset purchase\n"
        "• `/upgrade_asset` - upgrade an existing asset\n"
        "• `/sell_asset` - sell an existing asset\n"
        "• `/econ_adjust` - adjust a balance by delta\n"
        "• `/econ_set_balance` - set a balance directly\n"
        "• `/econ_set_kingdom_tax` - set a kingdom income tax\n"
        "• `/set_kingdom_treasury` - set a kingdom treasury amount\n"
        "• `/econ_grant_all` - grant all registered characters Val\n"
        "• `/transfer_val` - transfer Val between characters\n"
        "• `/assign-noble-title` - assign a noble title\n"
        "• `/upgrade-noble-title` - upgrade a noble title\n"
        "• `/edit-noble-title` - edit noble title presentation\n"
        "• `/econ_refresh_bank` - rebuild the bank dashboard\n"
    )
    await interaction.followup.send(msg, ephemeral=True)

@tree.command(name="econ_set_kingdom_tax", description="Set a kingdom's income tax rate (10–50%).", guild=discord.Object(id=GUILD_ID))
@staff_only()
@app_commands.describe(kingdom="Kingdom name (must match character/asset kingdom values).", rate="Tax rate percent.")
@app_commands.choices(
    kingdom=[
        app_commands.Choice(name="Sethrathiel", value="Sethrathiel"),
        app_commands.Choice(name="Velarith", value="Velarith"),
        app_commands.Choice(name="Lyvik", value="Lyvik"),
        app_commands.Choice(name="Baelon", value="Baelon"),
        app_commands.Choice(name="Avalea", value="Avalea"),
    ],
    rate=[
        app_commands.Choice(name="10%", value=10),
        app_commands.Choice(name="20%", value=20),
        app_commands.Choice(name="30%", value=30),
        app_commands.Choice(name="40%", value=40),
        app_commands.Choice(name="50%", value=50),
    ],
)
async def cmd_set_kingdom_tax(interaction: discord.Interaction, kingdom: str, rate: app_commands.Choice[int]):
    await interaction.response.defer(ephemeral=True)
    k = (kingdom or "").strip()
    if not k:
        await interaction.followup.send("Kingdom name is required.", ephemeral=True)
        return
    pct = int(rate.value)
    bp = pct * 100  # convert percent to basis points
    await upsert_kingdom_tax_bp(k, bp)
    await log_econ(interaction, "set_kingdom_tax", {"kingdom": k, "percent": pct, "tax_rate_bp": bp})
    trigger_bank_refresh()

    await interaction.followup.send(f"Set **{k}** tax rate to **{pct}%** (stored as **{bp} bp**).", ephemeral=True)



@tree.command(name="econ_adjust", description="(Staff) Adjust a character balance by delta.", guild=discord.Object(id=GUILD_ID))
@staff_only()
@app_commands.describe(character="Character name", delta="Positive or negative amount")
@app_commands.autocomplete(character=character_autocomplete)
async def cmd_econ_adjust(interaction: discord.Interaction, character: str, delta: int):
    await interaction.response.defer(ephemeral=True)

    delta = int(delta)
    cur_bal = await get_balance(character)
    proposed = int(cur_bal) + int(delta)

    if proposed < 0:
        await interaction.followup.send(
            "Denied: that adjustment would take the balance negative.\n"
            f"Available funds: **{format_currency(cur_bal)}**\n"
            f"Attempted adjustment: **{format_currency(delta)}**\n"
            f"Would result in: **{format_currency(proposed)}**",
            ephemeral=True,
        )
        return

    new_bal = await adjust_balance(character, delta)
    await log_econ(interaction, "adjust_balance", {"character": character, "delta": delta, "new_balance": new_bal})
    await interaction.followup.send(
        f"Adjusted **{character}** by **{format_currency(delta)}**. New balance: **{format_currency(new_bal)}**",
        ephemeral=True,
    )



@tree.command(name="set_kingdom_treasury", description="(Staff) Set a kingdom treasury to an exact value (Val).", guild=discord.Object(id=GUILD_ID))
@staff_only()
@app_commands.describe(kingdom="Kingdom name.", amount="New treasury amount (Val, must be >= 0).")
@app_commands.choices(
    kingdom=[
        app_commands.Choice(name="Sethrathiel", value="Sethrathiel"),
        app_commands.Choice(name="Velarith", value="Velarith"),
        app_commands.Choice(name="Lyvik", value="Lyvik"),
        app_commands.Choice(name="Baelon", value="Baelon"),
        app_commands.Choice(name="Avalea", value="Avalea"),
    ]
)
async def cmd_set_kingdom_treasury(interaction: discord.Interaction, kingdom: str, amount: int):
    await interaction.response.defer(ephemeral=True)
    k = (kingdom or "").strip()
    if not k:
        await interaction.followup.send("Kingdom name is required.", ephemeral=True)
        return
    amount = int(amount)
    if amount < 0:
        await interaction.followup.send("Amount must be >= 0.", ephemeral=True)
        return

    await set_kingdom_treasury_exact(k, amount)
    await log_econ(interaction, "set_kingdom_treasury", {"kingdom": k, "treasury": amount})
    trigger_bank_refresh()

    await interaction.followup.send(f"Set **{k}** treasury to **{format_currency(amount)}**.", ephemeral=True)
    trigger_bank_refresh()


@tree.command(name="econ_set_balance", description="(Staff) Set a character balance to an exact value.", guild=discord.Object(id=GUILD_ID))
@staff_only()
@app_commands.describe(character="Character name", value="New balance (must be >= 0)")
@app_commands.autocomplete(character=character_autocomplete)
async def cmd_econ_set_balance(interaction: discord.Interaction, character: str, value: int):
    await interaction.response.defer(ephemeral=True)

    value = int(value)
    if value < 0:
        cur_bal = await get_balance(character)
        await interaction.followup.send(
            "Denied: balance cannot be set to a negative value.\n"
            f"Current balance: **{format_currency(cur_bal)}**\n"
            f"Attempted set value: **{format_currency(value)}**",
            ephemeral=True,
        )
        return

    await set_balance(character, value)
    await log_econ(interaction, "set_balance", {"character": character, "value": value})
    await interaction.followup.send(f"Set **{character}** balance to **{format_currency(value)}**.", ephemeral=True)
    trigger_bank_refresh()


@tree.command(name="econ_grant_all", description="(Staff) Give every registered character an amount of Val.", guild=discord.Object(id=GUILD_ID))
@staff_only()
@app_commands.describe(amount="Amount to add to every character balance (Val)")
async def cmd_econ_grant_all(interaction: discord.Interaction, amount: int):
    """Grant all registered (non-archived) characters a flat Val amount.

    This is a bulk operation that affects econ_balances only. It does not touch assets,
    treasuries, taxes, or leaderboards.
    """
    await interaction.response.defer(ephemeral=True)

    try:
        amt = int(amount)
    except Exception:
        await interaction.followup.send("Amount must be an integer Val value.", ephemeral=True)
        return

    if amt == 0:
        await interaction.followup.send("Amount is 0 Val — no changes made.", ephemeral=True)
        return

    # Count eligible characters for reporting
    count_row = await db_fetchrow(
        """
        SELECT COUNT(*) AS n
        FROM characters
        WHERE guild_id=$1 AND archived=FALSE;
        """,
        DATA_GUILD_ID,
    )
    n_chars = int(count_row["n"]) if count_row else 0
    if n_chars <= 0:
        await interaction.followup.send("No registered characters found.", ephemeral=True)
        return

    status = await db_exec(
        """
        INSERT INTO econ_balances (guild_id, character_name, balance_val, updated_at)
        SELECT $1, c.name, $2::bigint, NOW()
        FROM characters c
        WHERE c.guild_id=$1 AND c.archived=FALSE
        ON CONFLICT (guild_id, character_name)
        DO UPDATE SET balance_val = econ_balances.balance_val + $2::bigint,
                      updated_at = NOW();
        """,
        DATA_GUILD_ID,
        amt,
    )

    await log_econ(interaction, "grant_all", {"amount": amt, "characters": n_chars, "status": status})
    trigger_bank_refresh()

    await interaction.followup.send(
        f"Granted **{format_currency(amt)}** to **{n_chars:,}** registered character(s).",
        ephemeral=True,
        allowed_mentions=discord.AllowedMentions.none(),
    )


@tree.command(name="transfer_val", description="Transfer Val from one of your characters to another character.", guild=discord.Object(id=GUILD_ID))
@app_commands.describe(
    source_character="One of your registered characters sending the funds",
    target_character="The registered character receiving the funds",
    amount="Amount of Val to transfer",
)
@app_commands.autocomplete(source_character=character_autocomplete, target_character=character_autocomplete)
async def cmd_transfer_val(interaction: discord.Interaction, source_character: str, target_character: str, amount: int):
    await interaction.response.defer(ephemeral=True)

    source_character = str(source_character or '').strip()
    target_character = str(target_character or '').strip()

    if not source_character or not target_character:
        await interaction.followup.send("Both source and target characters are required.", ephemeral=True)
        return

    if source_character == target_character:
        await interaction.followup.send("Source and target characters must be different.", ephemeral=True)
        return

    try:
        amount = int(amount)
    except Exception:
        await interaction.followup.send("Amount must be a whole number of Val.", ephemeral=True)
        return

    if amount <= 0:
        await interaction.followup.send("Transfer amount must be greater than 0 Val.", ephemeral=True)
        return

    source_owner = await get_character_owner(source_character)
    if source_owner is None:
        await interaction.followup.send("Source character not found.", ephemeral=True)
        return

    if int(source_owner) != int(interaction.user.id):
        await interaction.followup.send("You may only transfer funds from a character you own.", ephemeral=True)
        return

    target_owner = await get_character_owner(target_character)
    if target_owner is None:
        await interaction.followup.send("Target character not found.", ephemeral=True)
        return

    pool = await db_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            await conn.execute(
                """
                INSERT INTO econ_balances (guild_id, character_name, balance_val, updated_at)
                VALUES ($1, $2, 0, NOW())
                ON CONFLICT (guild_id, character_name) DO NOTHING;
                """,
                DATA_GUILD_ID,
                source_character,
            )
            await conn.execute(
                """
                INSERT INTO econ_balances (guild_id, character_name, balance_val, updated_at)
                VALUES ($1, $2, 0, NOW())
                ON CONFLICT (guild_id, character_name) DO NOTHING;
                """,
                DATA_GUILD_ID,
                target_character,
            )

            src_row = await conn.fetchrow(
                """
                SELECT balance_val
                FROM econ_balances
                WHERE guild_id=$1 AND character_name=$2
                FOR UPDATE;
                """,
                DATA_GUILD_ID,
                source_character,
            )
            tgt_row = await conn.fetchrow(
                """
                SELECT balance_val
                FROM econ_balances
                WHERE guild_id=$1 AND character_name=$2
                FOR UPDATE;
                """,
                DATA_GUILD_ID,
                target_character,
            )

            src_bal = int((src_row or {}).get('balance_val', 0) or 0)
            tgt_bal = int((tgt_row or {}).get('balance_val', 0) or 0)

            if src_bal < amount:
                await interaction.followup.send(
                    f"Insufficient funds. **{source_character}** has **{format_currency(src_bal)}**.",
                    ephemeral=True,
                )
                return

            new_src = src_bal - amount
            new_tgt = tgt_bal + amount

            await conn.execute(
                """
                UPDATE econ_balances
                SET balance_val=$1, updated_at=NOW()
                WHERE guild_id=$2 AND character_name=$3;
                """,
                new_src,
                DATA_GUILD_ID,
                source_character,
            )
            await conn.execute(
                """
                UPDATE econ_balances
                SET balance_val=$1, updated_at=NOW()
                WHERE guild_id=$2 AND character_name=$3;
                """,
                new_tgt,
                DATA_GUILD_ID,
                target_character,
            )

    await log_econ(
        interaction,
        "transfer_val",
        {
            "source_character": source_character,
            "target_character": target_character,
            "amount": amount,
            "source_new_balance": new_src,
            "target_new_balance": new_tgt,
        },
    )

    trigger_bank_refresh()

    await interaction.followup.send(
        f"Transferred **{format_currency(amount)}** from **{source_character}** to **{target_character}**.\n"
        f"**{source_character}**: **{format_currency(new_src)}**\n"
        f"**{target_character}**: **{format_currency(new_tgt)}**",
        ephemeral=True,
        allowed_mentions=discord.AllowedMentions.none(),
    )


@tree.command(name="econ_refresh_bank", description="(Staff) Refresh the bank dashboard messages.", guild=discord.Object(id=GUILD_ID))
@staff_only()
async def cmd_refresh_bank(interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)
    # Manual refresh (staff): force delete + recreate pages so formatting/content changes always appear.
    try:
        await force_rebuild_bank_dashboard()
        await interaction.followup.send("Bank dashboard rebuilt.", ephemeral=True)
    except Exception as e:
        await interaction.followup.send(f"Refresh failed: {e}", ephemeral=True)

@tree.command(name="purchase_new", description="(Staff) Record an asset purchase for a character.", guild=discord.Object(id=GUILD_ID))
@staff_only()
@app_commands.describe(
    character="Character purchasing the asset",
    asset_type="Asset category",
    tier="Tier being purchased",
    asset_name="Unique asset name (entered by staff)",
    kingdom="Kingdom that receives purchase/upgrade funds for this asset (required)",
)
@app_commands.choices(
    kingdom=[app_commands.Choice(name=k, value=k) for k in CANON_KINGDOMS]
)
@app_commands.autocomplete(character=character_autocomplete, asset_type=asset_type_autocomplete, tier=tier_autocomplete)
async def cmd_purchase_new(interaction: discord.Interaction, character: str, asset_type: str, tier: str, asset_name: str, kingdom: str):
    # Always defer quickly to avoid Discord timeouts
    await interaction.response.defer(ephemeral=True)

    owner = await get_character_owner(character)
    if owner is None:
        await interaction.followup.send("Character not found in DB.", ephemeral=True)
        return

    if is_noble_title_asset_type(asset_type):
        await interaction.followup.send("Noble Titles are staff-assigned only. Use `/assign-noble-title`.", ephemeral=True)
        return

    # Validate asset definition exists (and get add_income for audit)
    adef = await get_asset_def(asset_type, tier)
    if not adef:
        await interaction.followup.send("Invalid asset type/tier (not found in asset definitions).", ephemeral=True)
        return
    _tier_cost_val, add_income_val = adef

    # Cost is cumulative across tiers up to the selected target tier.
    cost_val = await cumulative_cost_to_tier(asset_type, tier)
    if cost_val is None:
        await interaction.followup.send("Unable to compute cumulative cost for this asset type/tier.", ephemeral=True)
        return

    # Kingdom is required for new asset purchases (no fallback/inference).
    sales_kingdom = str(kingdom or "").strip()
    if sales_kingdom not in CANON_KINGDOMS:
        await interaction.followup.send("Invalid kingdom selection.", ephemeral=True)
        return

    cur_bal = await get_balance(character)
    if cur_bal < cost_val:
        await interaction.followup.send(
            f"Insufficient funds. Balance **{format_currency(cur_bal)}**, cost **{format_currency(cost_val)}**.",
            ephemeral=True,
        )
        return

    asset_name = (asset_name or "").strip()
    if not asset_name:
        await interaction.followup.send("Asset name cannot be empty.", ephemeral=True)
        return

    # Allow same asset_name across different asset_type/tier, but not duplicates within the same type+tier.
    exists = await db_fetchrow(
        """
        SELECT 1
        FROM econ_assets
        WHERE guild_id=$1 AND character_name=$2 AND asset_type=$3 AND tier=$4 AND asset_name=$5
        LIMIT 1;
        """,
        DATA_GUILD_ID,
        character,
        asset_type,
        tier,
        asset_name,
    )
    if exists:
        await interaction.followup.send(
            "That character already has an asset with the same **type, tier, and name**. Choose a different name or tier.",
            ephemeral=True,
        )
        return

    # Record asset
    try:
        await db_exec(
            """
            INSERT INTO econ_assets (guild_id, character_name, user_id, asset_name, asset_type, tier, kingdom, created_at, updated_at)
            VALUES ($1, $2, $3, $4, $5, $6, $7, NOW(), NOW());
            """,
            DATA_GUILD_ID,
            character,
            int(owner),
            asset_name,
            asset_type,
            tier,
            sales_kingdom,
        )
    except Exception as e:
        await interaction.followup.send(f"Failed to add asset (see logs for details): {e}", ephemeral=True)
        return

    # Deduct cost
    new_bal = await adjust_balance(character, -cost_val)
    if sales_kingdom:
        await add_to_kingdom_treasury(sales_kingdom, int(cost_val))

    # Income is computed dynamically from assets; we don't store a separate total.
    new_daily_income = await recompute_daily_income(character)

    await log_econ(
        interaction,
        "purchase_new",
        {
            "character": character,
            "owner_user_id": int(owner),
            "asset_type": asset_type,
            "tier": tier,
            "asset_name": asset_name,
            "cost": cost_val,
            "add_income": add_income_val,
            "sales_kingdom": sales_kingdom,
            "new_balance": new_bal,
            "new_daily_income": new_daily_income,
        },
    )

    await interaction.followup.send(
        f"Recorded purchase for **{character}**:\n"
        f"• **{asset_type}** | **{tier}** | **{asset_name}**\n"
        f"Cost: **{format_currency(cost_val)}** → sent to **{sales_kingdom or 'N/A'}** treasury (new balance **{format_currency(new_bal)}**)\n"
        f"Daily income now: **{format_currency(new_daily_income)}**",
        ephemeral=True,
    )

@tree.command(name="upgrade_asset", description="(Staff) Upgrade an existing asset to a higher tier.", guild=discord.Object(id=GUILD_ID))
@staff_only()
@app_commands.autocomplete(character=character_autocomplete, asset=ac_asset_for_character, target_tier=ac_target_tier)
async def cmd_upgrade_asset(interaction: discord.Interaction, character: str, asset: str, target_tier: str):
    await interaction.response.defer(ephemeral=True)

    parts = [p.strip() for p in str(asset).split("|")]
    if len(parts) < 3:
        await interaction.followup.send("Invalid asset selection.", ephemeral=True)
        return
    asset_type = parts[0]
    current_tier = parts[1]
    asset_name = "|".join(parts[2:]).strip()

    if is_noble_title_asset_type(asset_type):
        await interaction.followup.send("Noble Titles must be upgraded with `/upgrade-noble-title`.", ephemeral=True)
        return

    exists = await db_fetchrow(
        '''
        SELECT 1
        FROM econ_assets
        WHERE guild_id=$1 AND character_name=$2 AND asset_type=$3 AND tier=$4 AND asset_name=$5
        LIMIT 1;
        ''',
        DATA_GUILD_ID,
        character,
        asset_type,
        current_tier,
        asset_name,
    )
    if not exists:
        await interaction.followup.send("That asset no longer exists on this character.", ephemeral=True)
        return

    cur_rank = _tier_rank(current_tier)
    tgt_rank = _tier_rank(target_tier)
    if cur_rank is not None and tgt_rank is not None and tgt_rank <= cur_rank:
        await interaction.followup.send("Target tier must be higher than current tier.", ephemeral=True)
        return

    cost_val = await incremental_cost_between_tiers(asset_type, current_tier, target_tier)
    if cost_val is None or cost_val <= 0:
        await interaction.followup.send("Unable to calculate upgrade cost for that tier change.", ephemeral=True)
        return

    # Destination kingdom for upgrade funds:
    # Prefer the asset's stored kingdom; if missing, fall back to definition kingdom for target tier; then character home kingdom.
    row_k = await db_fetchrow(
        '''
        SELECT COALESCE(kingdom, '') AS k
        FROM econ_assets
        WHERE guild_id=$1 AND character_name=$2 AND asset_type=$3 AND tier=$4 AND asset_name=$5
        LIMIT 1;
        ''',
        DATA_GUILD_ID,
        character,
        asset_type,
        current_tier,
        asset_name,
    )
    upgrade_kingdom = str((row_k or {}).get('k', '') or '').strip()
    if not upgrade_kingdom:
        upgrade_kingdom = await get_asset_definition_kingdom(asset_type, target_tier) or ''
    if not upgrade_kingdom:
        upgrade_kingdom = await get_character_kingdom(character) or ''

    cur_bal = await get_balance(character)
    if cur_bal < cost_val:
        await interaction.followup.send(
            f"Insufficient funds. Available: **{format_currency(cur_bal)}**. Required: **{format_currency(cost_val)}**.",
            ephemeral=True,
        )
        return

    await adjust_balance(character, -int(cost_val))
    if upgrade_kingdom:
        await add_to_kingdom_treasury(upgrade_kingdom, int(cost_val))
    await db_exec(
        '''
        UPDATE econ_assets
        SET tier=$1, kingdom=CASE WHEN COALESCE(kingdom,'')='' THEN $2 ELSE kingdom END
        WHERE guild_id=$3 AND character_name=$4 AND asset_type=$5 AND tier=$6 AND asset_name=$7;
        ''',
        target_tier,
        upgrade_kingdom,
        DATA_GUILD_ID,
        character,
        asset_type,
        current_tier,
        asset_name,
    )

    await log_econ(
        interaction,
        "upgrade_asset",
        {
            "character": character,
            "asset_type": asset_type,
            "asset_name": asset_name,
            "from_tier": current_tier,
            "to_tier": target_tier,
            "cost": int(cost_val),
            "sales_kingdom": upgrade_kingdom,
        },
    )

    trigger_bank_refresh()

    await interaction.followup.send(
        (
            f"Upgraded **{character}** asset:\n"
            f"- {asset_type} | {current_tier} | {asset_name}\n"
            f"→ {asset_type} | {target_tier} | {asset_name}\n"
            f"Cost: **{format_currency(cost_val)}** → sent to **{upgrade_kingdom or 'N/A'}** treasury\n"
            f"New balance: **{format_currency(await get_balance(character))}**"
        ),
        ephemeral=True,
    )



@tree.command(name="assign-noble-title", description="(Staff) Assign a noble title to a character.", guild=discord.Object(id=GUILD_ID))
@staff_only()
@app_commands.describe(
    character="Character receiving the noble title",
    title_name="Title family/rank",
    title_name_option="Exact displayed title form",
    kingdom="Kingdom associated with the title",
    custom_realm_name="Custom realm name for the title (leave blank for King/Queen/Regent/Sovereign)",
)
@app_commands.choices(
    kingdom=[app_commands.Choice(name=k, value=k) for k in CANON_KINGDOMS]
)
@app_commands.autocomplete(
    character=character_autocomplete,
    title_name=noble_title_family_autocomplete,
    title_name_option=noble_title_option_autocomplete,
)
async def cmd_assign_noble_title(interaction: discord.Interaction, character: str, title_name: str, title_name_option: str, kingdom: str, custom_realm_name: str = ""):
    await interaction.response.defer(ephemeral=True)

    owner = await get_character_owner(character)
    if owner is None:
        await interaction.followup.send("Character not found in DB.", ephemeral=True)
        return

    family = str(title_name or "").strip()
    if family not in NOBLE_TITLE_FAMILIES:
        await interaction.followup.send("Invalid noble title family selection.", ephemeral=True)
        return

    option = str(title_name_option or "").strip()
    if option not in NOBLE_TITLE_FAMILIES[family]["options"]:
        await interaction.followup.send("Invalid title name option for that family.", ephemeral=True)
        return

    title_kingdom = str(kingdom or "").strip()
    if title_kingdom not in CANON_KINGDOMS:
        await interaction.followup.send("Invalid kingdom selection.", ephemeral=True)
        return

    requires_realm = noble_title_family_requires_realm(family)
    realm_name = sanitize_plain_text(str(custom_realm_name or "")).strip()
    if requires_realm and not realm_name:
        await interaction.followup.send("Custom realm name cannot be empty for that title.", ephemeral=True)
        return
    if not requires_realm:
        realm_name = ""

    existing_same_slot = await db_fetchrow(
        """
        SELECT 1
        FROM econ_assets
        WHERE guild_id=$1 AND character_name=$2 AND asset_type='Noble Title'
          AND COALESCE(noble_realm_name,'')=$3 AND COALESCE(kingdom,'')=$4
        LIMIT 1;
        """,
        DATA_GUILD_ID,
        character,
        realm_name,
        title_kingdom,
    )
    if existing_same_slot:
        await interaction.followup.send(
            "That character already has a noble title for that realm/kingdom slot. Use `/upgrade-noble-title` or `/edit-noble-title` instead.",
            ephemeral=True,
        )
        return

    tier = str(NOBLE_TITLE_FAMILIES[family]["tier"])
    cost_val = int(NOBLE_TITLE_FAMILIES[family]["cost"])
    cur_bal = await get_balance(character)
    if cost_val > 0 and cur_bal < cost_val:
        await interaction.followup.send(
            f"Insufficient funds. Balance **{format_currency(cur_bal)}**, cost **{format_currency(cost_val)}**.",
            ephemeral=True,
        )
        return

    await db_exec(
        """
        INSERT INTO econ_assets (
            guild_id, character_name, user_id, asset_name, asset_type, tier, kingdom,
            noble_title_family, noble_title_option, noble_realm_name,
            created_at, updated_at
        )
        VALUES ($1, $2, $3, $4, 'Noble Title', $5, $6, $7, $8, $9, NOW(), NOW());
        """,
        DATA_GUILD_ID,
        character,
        int(owner),
        option,
        tier,
        title_kingdom,
        family,
        option,
        realm_name,
    )

    new_bal = cur_bal
    if cost_val > 0:
        new_bal = await adjust_balance(character, -cost_val)
        await add_to_kingdom_treasury(title_kingdom, cost_val)

    await log_econ(
        interaction,
        "assign_noble_title",
        {
            "character": character,
            "title_family": family,
            "title_option": option,
            "realm_name": realm_name,
            "kingdom": title_kingdom,
            "cost": cost_val,
            "new_balance": new_bal,
        },
    )

    trigger_bank_refresh()

    title_display = noble_title_display_from_parts(option, realm_name, title_kingdom, family)
    cost_msg = f" Cost: **{format_currency(cost_val)}**." if cost_val > 0 else " No cost."
    await interaction.followup.send(
        f"Assigned **{title_display}** to **{character}**.{cost_msg}",
        ephemeral=True,
    )


@tree.command(name="upgrade-noble-title", description="(Staff) Upgrade an existing noble title.", guild=discord.Object(id=GUILD_ID))
@staff_only()
@app_commands.describe(
    character="Character whose noble title is being upgraded",
    current_title="Existing noble title to upgrade",
    new_title_name="New title family/rank",
    new_title_name_option="Exact displayed title form for the new rank",
)
@app_commands.autocomplete(
    character=character_autocomplete,
    current_title=noble_title_realm_autocomplete,
    new_title_name=noble_title_family_autocomplete,
    new_title_name_option=noble_title_option_autocomplete,
)
async def cmd_upgrade_noble_title(interaction: discord.Interaction, character: str, current_title: str, new_title_name: str, new_title_name_option: str):
    await interaction.response.defer(ephemeral=True)

    cur_option_lookup, realm_name_lookup, title_kingdom_lookup = parse_noble_title_identifier(current_title)
    if not cur_option_lookup:
        await interaction.followup.send("Select an existing noble title to upgrade.", ephemeral=True)
        return

    row = await db_fetchrow(
        """
        SELECT COALESCE(noble_title_family,'') AS noble_title_family,
               COALESCE(noble_title_option,'') AS noble_title_option,
               COALESCE(noble_realm_name,'') AS noble_realm_name,
               COALESCE(kingdom,'') AS kingdom,
               tier,
               asset_name
        FROM econ_assets
        WHERE guild_id=$1 AND character_name=$2 AND asset_type='Noble Title'
          AND COALESCE(noble_title_option,'')=$3 AND COALESCE(noble_realm_name,'')=$4 AND COALESCE(kingdom,'')=$5
        LIMIT 1;
        """,
        DATA_GUILD_ID,
        character,
        cur_option_lookup,
        realm_name_lookup,
        title_kingdom_lookup,
    )
    if not row:
        await interaction.followup.send("That noble title was not found on this character.", ephemeral=True)
        return

    new_family = str(new_title_name or "").strip()
    if new_family not in NOBLE_TITLE_FAMILIES:
        await interaction.followup.send("Invalid target noble title family.", ephemeral=True)
        return

    new_option = str(new_title_name_option or "").strip()
    if new_option not in NOBLE_TITLE_FAMILIES[new_family]["options"]:
        await interaction.followup.send("Invalid title option for that family.", ephemeral=True)
        return

    cur_family = str(row.get("noble_title_family") or TIER_TO_NOBLE_TITLE_FAMILY.get(str(row.get("tier") or "")) or "").strip()
    cur_option = str(row.get("noble_title_option") or row.get("asset_name") or "").strip()
    realm_name = str(row.get("noble_realm_name") or "").strip()
    title_kingdom = str(row.get("kingdom") or "").strip()

    cur_rank = noble_title_rank_for_family(cur_family)
    new_rank = noble_title_rank_for_family(new_family)
    if cur_rank and new_rank and new_rank < cur_rank:
        await interaction.followup.send("Noble titles cannot be downgraded with this command.", ephemeral=True)
        return

    if noble_title_family_requires_realm(new_family) != noble_title_family_requires_realm(cur_family):
        await interaction.followup.send("Use `/edit-noble-title` to switch between realm titles and sovereign titles.", ephemeral=True)
        return

    current_cost = noble_title_cost_for_family(cur_family)
    target_cost = noble_title_cost_for_family(new_family)
    upgrade_cost = max(0, int(target_cost) - int(current_cost))

    cur_bal = await get_balance(character)
    if upgrade_cost > 0 and cur_bal < upgrade_cost:
        await interaction.followup.send(
            f"Insufficient funds. Balance **{format_currency(cur_bal)}**, upgrade cost **{format_currency(upgrade_cost)}**.",
            ephemeral=True,
        )
        return

    await db_exec(
        """
        UPDATE econ_assets
        SET tier=$1,
            asset_name=$2,
            noble_title_family=$3,
            noble_title_option=$4,
            updated_at=NOW()
        WHERE guild_id=$5 AND character_name=$6 AND asset_type='Noble Title'
          AND COALESCE(noble_title_option,'')=$7 AND COALESCE(noble_realm_name,'')=$8 AND COALESCE(kingdom,'')=$9;
        """,
        str(NOBLE_TITLE_FAMILIES[new_family]["tier"]),
        new_option,
        new_family,
        new_option,
        DATA_GUILD_ID,
        character,
        cur_option_lookup,
        realm_name_lookup,
        title_kingdom_lookup,
    )

    new_bal = cur_bal
    if upgrade_cost > 0:
        new_bal = await adjust_balance(character, -upgrade_cost)
        if title_kingdom:
            await add_to_kingdom_treasury(title_kingdom, upgrade_cost)

    await log_econ(
        interaction,
        "upgrade_noble_title",
        {
            "character": character,
            "realm_name": realm_name,
            "from_title_family": cur_family,
            "from_title_option": cur_option,
            "to_title_family": new_family,
            "to_title_option": new_option,
            "kingdom": title_kingdom,
            "cost": upgrade_cost,
            "new_balance": new_bal,
        },
    )

    trigger_bank_refresh()

    before_display = noble_title_display_from_parts(cur_option, realm_name, title_kingdom, cur_family)
    after_display = noble_title_display_from_parts(new_option, realm_name, title_kingdom, new_family)
    await interaction.followup.send(
        f"Updated noble title for **{character}**: **{before_display}** → **{after_display}**. Upgrade cost: **{format_currency(upgrade_cost)}**.",
        ephemeral=True,
    )


@tree.command(name="edit-noble-title", description="(Staff) Correct an existing noble title.", guild=discord.Object(id=GUILD_ID))
@staff_only()
@app_commands.describe(
    character="Character whose noble title is being corrected",
    current_title="Existing noble title to edit",
    new_title_name="Corrected title family/rank",
    new_title_name_option="Corrected displayed title form",
    kingdom="Corrected kingdom",
    custom_realm_name="Corrected custom realm name (leave blank for King/Queen/Regent/Sovereign)",
)
@app_commands.choices(
    kingdom=[app_commands.Choice(name=k, value=k) for k in CANON_KINGDOMS]
)
@app_commands.autocomplete(
    character=character_autocomplete,
    current_title=noble_title_realm_autocomplete,
    new_title_name=noble_title_family_autocomplete,
    new_title_name_option=noble_title_option_autocomplete,
)
async def cmd_edit_noble_title(interaction: discord.Interaction, character: str, current_title: str, new_title_name: str, new_title_name_option: str, kingdom: str, custom_realm_name: str = ""):
    await interaction.response.defer(ephemeral=True)

    old_option, old_realm, old_kingdom = parse_noble_title_identifier(current_title)
    if not old_option:
        await interaction.followup.send("Select an existing noble title to edit.", ephemeral=True)
        return

    row = await db_fetchrow(
        """
        SELECT COALESCE(noble_title_family,'') AS noble_title_family,
               COALESCE(noble_title_option,'') AS noble_title_option,
               COALESCE(noble_realm_name,'') AS noble_realm_name,
               COALESCE(kingdom,'') AS kingdom,
               tier,
               asset_name
        FROM econ_assets
        WHERE guild_id=$1 AND character_name=$2 AND asset_type='Noble Title'
          AND COALESCE(noble_title_option,'')=$3 AND COALESCE(noble_realm_name,'')=$4 AND COALESCE(kingdom,'')=$5
        LIMIT 1;
        """,
        DATA_GUILD_ID,
        character,
        old_option,
        old_realm,
        old_kingdom,
    )
    if not row:
        await interaction.followup.send("That noble title was not found on this character.", ephemeral=True)
        return

    new_family = str(new_title_name or "").strip()
    if new_family not in NOBLE_TITLE_FAMILIES:
        await interaction.followup.send("Invalid noble title family selection.", ephemeral=True)
        return

    new_option = str(new_title_name_option or "").strip()
    if new_option not in NOBLE_TITLE_FAMILIES[new_family]["options"]:
        await interaction.followup.send("Invalid title name option for that family.", ephemeral=True)
        return

    new_kingdom = str(kingdom or "").strip()
    if new_kingdom not in CANON_KINGDOMS:
        await interaction.followup.send("Invalid kingdom selection.", ephemeral=True)
        return

    requires_realm = noble_title_family_requires_realm(new_family)
    new_realm = sanitize_plain_text(str(custom_realm_name or "")).strip()
    if requires_realm and not new_realm:
        await interaction.followup.send("Custom realm name cannot be empty for that title.", ephemeral=True)
        return
    if not requires_realm:
        new_realm = ""

    conflict = await db_fetchrow(
        """
        SELECT 1
        FROM econ_assets
        WHERE guild_id=$1 AND character_name=$2 AND asset_type='Noble Title'
          AND COALESCE(noble_title_option,'')=$3 AND COALESCE(noble_realm_name,'')=$4 AND COALESCE(kingdom,'')=$5
          AND NOT (COALESCE(noble_title_option,'')=$6 AND COALESCE(noble_realm_name,'')=$7 AND COALESCE(kingdom,'')=$8)
        LIMIT 1;
        """,
        DATA_GUILD_ID,
        character,
        new_option,
        new_realm,
        new_kingdom,
        old_option,
        old_realm,
        old_kingdom,
    )
    if conflict:
        await interaction.followup.send("That corrected noble title would duplicate another title on this character.", ephemeral=True)
        return

    await db_exec(
        """
        UPDATE econ_assets
        SET tier=$1,
            asset_name=$2,
            kingdom=$3,
            noble_title_family=$4,
            noble_title_option=$5,
            noble_realm_name=$6,
            updated_at=NOW()
        WHERE guild_id=$7 AND character_name=$8 AND asset_type='Noble Title'
          AND COALESCE(noble_title_option,'')=$9 AND COALESCE(noble_realm_name,'')=$10 AND COALESCE(kingdom,'')=$11;
        """,
        str(NOBLE_TITLE_FAMILIES[new_family]["tier"]),
        new_option,
        new_kingdom,
        new_family,
        new_option,
        new_realm,
        DATA_GUILD_ID,
        character,
        old_option,
        old_realm,
        old_kingdom,
    )

    old_family = str(row.get("noble_title_family") or TIER_TO_NOBLE_TITLE_FAMILY.get(str(row.get("tier") or "")) or "").strip()
    old_display = noble_title_display_from_parts(old_option, old_realm, old_kingdom, old_family)
    new_display = noble_title_display_from_parts(new_option, new_realm, new_kingdom, new_family)

    await log_econ(
        interaction,
        "edit_noble_title",
        {
            "character": character,
            "from_title": old_display,
            "to_title": new_display,
        },
    )

    trigger_bank_refresh()

    await interaction.followup.send(
        f"Corrected noble title for **{character}**: **{old_display}** → **{new_display}**.",
        ephemeral=True,
    )


@tree.command(name="sell_asset", description="(Staff) Sell an owned asset and refund 100% cumulative invested Val (taxed by asset kingdom).", guild=discord.Object(id=GUILD_ID))
@staff_only()
@app_commands.autocomplete(character=character_autocomplete, asset=ac_asset_for_character)
async def cmd_sell_asset(interaction: discord.Interaction, character: str, asset: str):
    await interaction.response.defer(ephemeral=True)

    parts = [p.strip() for p in str(asset).split("|")]
    if len(parts) < 3:
        await interaction.followup.send("Invalid asset selection.", ephemeral=True)
        return
    asset_type = parts[0]
    tier = parts[1]
    asset_name = "|".join(parts[2:]).strip()

    if is_noble_title_asset_type(asset_type):
        await interaction.followup.send("Noble Titles cannot be sold. Remove them manually if needed.", ephemeral=True)
        return

    row = await db_fetchrow(
        '''
        SELECT COALESCE(kingdom,'') AS kingdom
        FROM econ_assets
        WHERE guild_id=$1 AND character_name=$2 AND asset_type=$3 AND tier=$4 AND asset_name=$5
        LIMIT 1;
        ''',
        DATA_GUILD_ID,
        character,
        asset_type,
        tier,
        asset_name,
    )
    if not row:
        await interaction.followup.send("That asset no longer exists on this character.", ephemeral=True)
        return

    asset_kingdom = str(row.get("kingdom") or "").strip()
    if not asset_kingdom:
        # Safety fallback for legacy rows; do not block sale.
        asset_kingdom = str((await get_character_kingdom(character)) or "").strip()

    cumulative_cost = await cumulative_cost_to_tier(asset_type, tier)
    if cumulative_cost is None or cumulative_cost <= 0:
        await interaction.followup.send("Unable to calculate cumulative cost for this asset.", ephemeral=True)
        return
    gross_refund = int(cumulative_cost)

    tax_bp = 0
    if gross_refund > 0 and asset_kingdom:
        tax_bp = int(await get_kingdom_tax_bp(asset_kingdom))
    tax_amount = int((gross_refund * tax_bp) // 10000) if gross_refund > 0 and tax_bp > 0 else 0
    net_refund = gross_refund - tax_amount

    await db_exec(
        '''
        DELETE FROM econ_assets
        WHERE guild_id=$1 AND character_name=$2 AND asset_type=$3 AND tier=$4 AND asset_name=$5;
        ''',
        DATA_GUILD_ID,
        character,
        asset_type,
        tier,
        asset_name,
    )

    if net_refund:
        await adjust_balance(character, int(net_refund))
    if tax_amount and asset_kingdom:
        await add_to_kingdom_treasury(asset_kingdom, int(tax_amount))

    await log_econ(
        interaction,
        "sell_asset",
        {
            "character": character,
            "asset_type": asset_type,
            "tier": tier,
            "asset_name": asset_name,
            "asset_kingdom": asset_kingdom,
            "gross_refund": gross_refund,
            "tax_bp": tax_bp,
            "tax_amount": tax_amount,
            "net_refund": net_refund,
        },
    )

    trigger_bank_refresh()

    msg = (
        f"Sold/removed asset from **{character}**:\n"
        f"- {asset_type} | {tier} | {asset_name}\n"
    )
    if gross_refund:
        msg += f"Gross refund: **{format_currency(gross_refund)}**\n"
        if tax_amount and asset_kingdom:
            msg += f"Tax to **{asset_kingdom}**: **{format_currency(tax_amount)}** ({tax_bp/100:.2f}%)\n"
        msg += f"Net to character: **{format_currency(net_refund)}**\n"
    msg += f"New balance: **{format_currency(await get_balance(character))}**"

    await interaction.followup.send(msg, ephemeral=True)
# -------------------------
# Global app command error handler (prevents "stuck thinking" on exceptions)
# -------------------------

@tree.error
async def on_app_command_error(interaction: discord.Interaction, error: app_commands.AppCommandError) -> None:
    # Always log the root error in Railway logs
    try:
        import traceback
        traceback.print_exception(type(error), error, error.__traceback__)
    except Exception:
        pass

    # Respond ephemerally so interactions don't hang forever
    msg = "⚠️ Internal error while running that command. Check Railway logs for details."
    try:
        if interaction.response.is_done():
            await interaction.followup.send(msg, ephemeral=True)
        else:
            await interaction.response.send_message(msg, ephemeral=True)
    except Exception:
        pass

# -------------------------
# Startup / sync
# -------------------------


async def delete_all_guild_commands():
    """Delete ALL guild-scoped commands in the configured GUILD_ID on Discord.
    This is a server-side cleanup to eliminate stale signatures/duplicates.
    """
    try:
        guild_obj = discord.Object(id=GUILD_ID)
        guild_cmds = await tree.fetch_commands(guild=guild_obj)
        if guild_cmds:
            for c in guild_cmds:
                try:
                    await c.delete()
                except Exception as e:
                    print(f"[warn] Failed deleting GUILD /{getattr(c,'name','?')}: {e}")
            print(f"[test] Requested deletion of {len(guild_cmds)} GUILD command(s).")
        else:
            print("[test] No GUILD commands found to delete.")
    except Exception as e:
        print(f"[warn] Guild command deletion failed/skipped: {e}")

async def delete_all_global_commands() -> None:
    # You selected option B: bot deletes global commands automatically.
    try:
        global_cmds = await tree.fetch_commands()  # global
        if global_cmds:
            for c in global_cmds:
                try:
                    await c.delete()
                except Exception as e:
                    print(f"[warn] Failed deleting GLOBAL /{getattr(c,'name','?')}: {e}")
            print(f"[test] Requested deletion of {len(global_cmds)} GLOBAL command(s).")
        else:
            print("[test] No GLOBAL commands found.")
    except Exception as e:
        print(f"[warn] Global command deletion failed/skipped: {e}")


async def _daily_reminder_already_sent(reminder_date: date) -> bool:
    row = await db_fetchrow(
        "SELECT 1 FROM econ_daily_reminder_runs WHERE guild_id=$1 AND reminder_date=$2 LIMIT 1;",
        DATA_GUILD_ID,
        reminder_date,
    )
    return bool(row)


async def _mark_daily_reminder_sent(reminder_date: date) -> None:
    await db_exec(
        """
        INSERT INTO econ_daily_reminder_runs (guild_id, reminder_date)
        VALUES ($1, $2)
        ON CONFLICT (guild_id, reminder_date) DO NOTHING;
        """,
        DATA_GUILD_ID,
        reminder_date,
    )


@tasks.loop(minutes=1)
async def daily_income_reminder_loop():
    now = datetime.now(CHICAGO_TZ)
    minutes_now = now.hour * 60 + now.minute
    target_minutes = DAILY_REMINDER_HOUR * 60 + DAILY_REMINDER_MINUTE
    if not (target_minutes <= minutes_now <= target_minutes + 2):
        return
    today = now.date()
    try:
        if await _daily_reminder_already_sent(today):
            return
    except Exception as e:
        print(f"[warn] Could not check daily reminder log: {e}")
        return

    guild = client.get_guild(GUILD_ID)
    if guild is None:
        try:
            guild = await client.fetch_guild(GUILD_ID)
        except Exception as e:
            print(f"[warn] Daily reminder guild fetch failed: {e}")
            return
    channel = guild.get_channel(DAILY_REMINDER_CHANNEL_ID) if hasattr(guild, 'get_channel') else None
    if channel is None:
        try:
            channel = await client.fetch_channel(DAILY_REMINDER_CHANNEL_ID)
        except Exception as e:
            print(f"[warn] Daily reminder channel fetch failed: {e}")
            return
    try:
        await channel.send(
            DAILY_REMINDER_MESSAGE,
            allowed_mentions=discord.AllowedMentions(roles=True, users=False, everyone=False),
        )
        await _mark_daily_reminder_sent(today)
        print(f"[test] Sent daily income reminder for {today.isoformat()} to channel {DAILY_REMINDER_CHANNEL_ID}")
    except Exception as e:
        print(f"[warn] Sending daily income reminder failed: {e}")


@daily_income_reminder_loop.before_loop
async def _before_daily_income_reminder_loop():
    await client.wait_until_ready()


@client.event
async def on_ready():
    # Keep on_ready resilient: never allow an exception to abort command sync.
    print(f"[test] Starting {APP_VERSION}…")
    print(f"[test] Logged in as {client.user} (commands guild: {GUILD_ID}; data guild: {DATA_GUILD_ID})")
    print(f"[debug] raw STAFF_ROLE_IDS env: {repr(_get('STAFF_ROLE_IDS',''))}")
    print(f"[debug] STAFF_ROLE_IDS_DEFAULT: {sorted(list(STAFF_ROLE_IDS_DEFAULT))}")
    print(f"[debug] STAFF_ROLE_IDS (effective): {sorted(list(STAFF_ROLE_IDS))}")

    guild_obj = discord.Object(id=GUILD_ID)

    # --- COMMAND SYNC (first, hardened) ---
    try:
        # Defensive: ensure new commands are present in local guild registry.
        # NOTE: decorated functions are Command objects.
        try:
            tree.add_command(cmd_upgrade_asset, guild=guild_obj)
        except Exception:
            pass
        try:
            tree.add_command(cmd_sell_asset, guild=guild_obj)
        except Exception:
            pass
        try:
            tree.add_command(cmd_assign_noble_title, guild=guild_obj)
        except Exception:
            pass
        try:
            tree.add_command(cmd_upgrade_noble_title, guild=guild_obj)
        except Exception:
            pass
        try:
            tree.add_command(cmd_edit_noble_title, guild=guild_obj)
        except Exception:
            pass

        # Optional: copy any locally-registered global commands into guild scope (no global sync).
        try:
            tree.copy_global_to(guild=guild_obj)
        except Exception as e:
            print(f"[warn] copy_global_to failed/skipped: {e}")

# NOTE: We do NOT clear guild commands here.
# tree.sync(guild=...) overwrites the server-side guild command set to match the locally-registered guild commands.
# Clearing first would drop all existing commands unless we re-register every single one manually.

        synced = await tree.sync(guild=guild_obj)
        print(f"[test] Synced {len(synced)} guild command(s).")

        # Post-sync verification: what Discord now has.
        try:
            post = sorted([c.name for c in await tree.fetch_commands(guild=guild_obj)])
            print(f"[debug] Post-sync guild commands (server): {post}")
        except Exception as e:
            print(f"[warn] Could not fetch post-sync guild commands: {e}")

        # Cleanup global commands last (best-effort) to avoid duplicates from prior versions.
        try:
            await delete_all_global_commands()
        except Exception as e:
            print(f"[warn] Global command deletion failed/skipped: {e}")

    except Exception as e:
        print(f"[warn] Command sync block failed: {e}")

    # --- DB / SEED / BANK (best-effort, after sync) ---
    try:
        await ensure_schema()
    except Exception as e:
        print(f"[warn] ensure_schema failed: {e}")

    try:
        await seed_asset_definitions()
    except Exception as e:
        print(f"[warn] seed_asset_definitions failed: {e}")
    try:
        if not daily_income_reminder_loop.is_running():
            daily_income_reminder_loop.start()
            print(f"[test] Started daily income reminder loop for {DAILY_REMINDER_HOUR:02d}:{DAILY_REMINDER_MINUTE:02d} America/Chicago")
    except Exception as e:
        print(f"[warn] Failed to start daily income reminder loop: {e}")


def main():
    client.run(DISCORD_TOKEN)


if __name__ == "__main__":
    main(

)
